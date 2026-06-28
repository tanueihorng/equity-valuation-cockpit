#!/usr/bin/env python3
"""
Equity Valuation Cockpit - production sell-side equity research model.
Builds a fully-linked Excel workbook with LIVE cell formulas implementing the
canonical financial engine. All inputs live on ASSUMPTIONS; every other sheet
references them so editing an input flows everywhere.

openpyxl 3.0.10
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import (
    ColorScaleRule, CellIsRule, FormulaRule
)
from openpyxl.comments import Comment
from openpyxl.chart import BarChart, ScatterChart, Reference, Series

# ----------------------------------------------------------------------------
# SHARED DESIGN SYSTEM  (hex without leading '#', ARGB for openpyxl)
# ----------------------------------------------------------------------------
C_BG        = "0A0E1A"   # page background near-black navy
C_PANEL     = "131A2B"   # panel / surface
C_RAISED    = "161E33"   # raised surface
C_HAIR      = "232C45"   # hairline borders
C_TEXT      = "E6ECF5"   # primary text
C_MUTED     = "8A96B0"   # muted text
C_GOLD      = "C9A227"   # gold accent (brand)
C_GOLD_BR   = "E5B83B"   # bright gold
C_GREEN     = "2DD4A7"   # positive
C_RED       = "F0556B"   # negative
C_BLUE      = "4C8DFF"   # info blue
C_AMBER     = "F2B33D"   # warning amber
C_INPUT     = "1C2742"   # input cell fill

# Fonts (numbers in monospace)
F_UI    = "Inter"
F_MONO  = "Menlo"   # monospace face that ships on macOS Excel

# ---- Fills ----
fill_bg     = PatternFill("solid", fgColor=C_BG)
fill_panel  = PatternFill("solid", fgColor=C_PANEL)
fill_raised = PatternFill("solid", fgColor=C_RAISED)
fill_input  = PatternFill("solid", fgColor=C_INPUT)
fill_gold   = PatternFill("solid", fgColor=C_GOLD)
fill_green  = PatternFill("solid", fgColor=C_GREEN)
fill_red    = PatternFill("solid", fgColor=C_RED)
fill_amber  = PatternFill("solid", fgColor=C_AMBER)
fill_teal   = PatternFill("solid", fgColor=C_GREEN)
fill_blue   = PatternFill("solid", fgColor=C_BLUE)

# ---- Borders ----
hair = Side(style="thin", color=C_HAIR)
gold_side = Side(style="medium", color=C_GOLD)
border_hair = Border(left=hair, right=hair, top=hair, bottom=hair)
border_input = Border(left=gold_side, right=hair, top=hair, bottom=hair)
border_bottom = Border(bottom=hair)

# ---- Number formats ----
NF_NUM   = "#,##0"
NF_NUM1  = "#,##0.0"
NF_PCT   = "0.0%"
NF_PCT2  = "0.00%"
NF_USD   = "$#,##0.00"
NF_USDM  = "$#,##0"
NF_MULT  = "0.00x"
NF_PX    = "$#,##0.00"
NF_SCORE = "0.00"

# ----------------------------------------------------------------------------
# Style helper functions
# ----------------------------------------------------------------------------

def style_cell(ws, ref, value=None, *, font_color=C_TEXT, size=10, bold=False,
               italic=False, mono=False, fill=None, align="left", valign="center",
               numfmt=None, border=None, wrap=False):
    c = ws[ref]
    if value is not None:
        c.value = value
    c.font = Font(name=(F_MONO if mono else F_UI), size=size, bold=bold,
                  italic=italic, color=font_color)
    if fill is not None:
        c.fill = fill
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if numfmt:
        c.number_format = numfmt
    if border is not None:
        c.border = border
    return c


def banner(ws, title, subtitle, last_col="K"):
    """Dark banner title spanning columns A..last_col across rows 1-3."""
    last = openpyxl.utils.column_index_from_string(last_col)
    # fill background for banner area
    for r in (1, 2, 3):
        for col in range(1, last + 1):
            cell = ws.cell(row=r, column=col)
            cell.fill = fill_raised
    ws.merge_cells(f"A1:{last_col}1")
    ws.merge_cells(f"A2:{last_col}2")
    ws.merge_cells(f"A3:{last_col}3")
    style_cell(ws, "A1", title, font_color=C_GOLD_BR, size=18, bold=True,
               fill=fill_raised, align="left")
    style_cell(ws, "A2", subtitle, font_color=C_MUTED, size=10, italic=True,
               fill=fill_raised, align="left")
    style_cell(ws, "A3", "EQUITY VALUATION COCKPIT  •  APEX TECHNOLOGIES (SAMPLE)  •  APEX",
               font_color=C_MUTED, size=8, fill=fill_raised, align="left")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 14


def section_label(ws, ref, text, fill=None):
    """Small ALL-CAPS letter-spaced muted section label."""
    spaced = "  ".join(list(text.upper()))
    style_cell(ws, ref, spaced, font_color=C_MUTED, size=8, bold=True,
               fill=(fill or fill_panel), align="left")


def paint_background(ws, last_col="N", last_row=120):
    """Paint the panel background over the working area for the dark look."""
    last = openpyxl.utils.column_index_from_string(last_col)
    for r in range(4, last_row + 1):
        for col in range(1, last + 1):
            cell = ws.cell(row=r, column=col)
            if cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = fill_panel


def lbl(ws, ref, text, **kw):
    kw.setdefault("font_color", C_MUTED)
    kw.setdefault("size", 9)
    style_cell(ws, ref, text, **kw)


def val(ws, ref, formula, numfmt=NF_NUM, *, color=C_TEXT, bold=False, size=10,
        fill=None, border=border_hair, align="right"):
    style_cell(ws, ref, formula, font_color=color, size=size, bold=bold,
               mono=True, fill=(fill or fill_panel), align=align, numfmt=numfmt,
               border=border)


def inp(ws, ref, value, numfmt=NF_NUM, note=None):
    """Editable input cell with distinct fill + gold left border."""
    c = style_cell(ws, ref, value, font_color=C_GOLD_BR, size=10, bold=True,
                   mono=True, fill=fill_input, align="right", numfmt=numfmt,
                   border=border_input)
    if note:
        cm = Comment(note, "Model")
        cm.width = 240
        cm.height = 110
        c.comment = cm
    return c


def setcols(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def no_grid_freeze(ws, freeze="A4"):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze
    ws.sheet_properties.tabColor = C_PANEL


# ============================================================================
# BUILD WORKBOOK
# ============================================================================
wb = Workbook()

# create sheets in desired tab order
ws_dash = wb.active
ws_dash.title = "DASHBOARD"
ws_assum = wb.create_sheet("ASSUMPTIONS")
ws_dcf   = wb.create_sheet("DCF")
ws_rel   = wb.create_sheet("RELATIVE")
ws_val   = wb.create_sheet("VALUE_HEURISTICS")
ws_qual  = wb.create_sheet("QUALITY")
ws_time  = wb.create_sheet("TIMING")
ws_mc    = wb.create_sheet("SCENARIO_MC")
ws_context = wb.create_sheet("CONTEXT")

# ============================================================================
# 2) ASSUMPTIONS  (build first so other sheets can reference named cells)
#    Inputs all in column C.  Notes in column E.
# ============================================================================
A = ws_assum
banner(A, "ASSUMPTIONS", "Single source of truth — every gold cell is editable; all other sheets reference these.", "F")
no_grid_freeze(A, "A4")
paint_background(A, "F", 130)
setcols(A, {"A": 3, "B": 34, "C": 16, "D": 3, "E": 60, "F": 3})

# We define inputs as (label, cellref, value, numfmt, note)
# Use an assignment-by-row builder.
row = 5

def grp(title, r):
    section_label(A, f"B{r}", title, fill=fill_raised)
    for col in ("A", "C", "D", "E", "F"):
        A[f"{col}{r}"].fill = fill_raised
    style_cell(A, f"C{r}", "VALUE", font_color=C_MUTED, size=8, bold=True,
               fill=fill_raised, align="right")
    style_cell(A, f"E{r}", "NOTE", font_color=C_MUTED, size=8, bold=True,
               fill=fill_raised, align="left")
    return r + 1

def put(r, label, value, numfmt, note):
    lbl(A, f"B{r}", label, size=10, font_color=C_TEXT)
    inp(A, f"C{r}", value, numfmt, note)
    style_cell(A, f"E{r}", note, font_color=C_MUTED, size=8, italic=True,
               fill=fill_panel, align="left", wrap=True)
    return r + 1

# ---- MARKET ----
row = grp("MARKET", row)
r_price   = row; row = put(row, "Current price ($/sh)", 298.01, NF_USD, "Live market price per share")
r_shares  = row; row = put(row, "Shares outstanding (M)", 14800, NF_NUM, "Diluted shares, millions")
r_hi52    = row; row = put(row, "52-week high ($)", 317.4, NF_USD, "52w high")
r_lo52    = row; row = put(row, "52-week low ($)", 195.07, NF_USD, "52w low")
r_swhi    = row; row = put(row, "Swing high ($)", 317.4, NF_USD, "Recent swing high for Fib")
r_swlo    = row; row = put(row, "Swing low ($)", 195.07, NF_USD, "Recent swing low for Fib")
r_ma50    = row; row = put(row, "50-day MA ($)", 302, NF_USD, "50DMA")
r_ma200   = row; row = put(row, "200-day MA ($)", 258, NF_USD, "200DMA")
r_rsi     = row; row = put(row, "RSI(14)", 54, NF_NUM1, "Relative strength index")
r_beta    = row; row = put(row, "Beta", 1.09, "0.00", "Levered equity beta")
row += 1

# ---- INCOME & CF (TTM) ----
row = grp("INCOME & CASH FLOW (TTM)", row)
r_rev     = row; row = put(row, "Revenue ($M)", 416160, NF_NUM, "TTM revenue, millions")
r_ebitm   = row; row = put(row, "EBIT margin", 0.32, NF_PCT, "Operating margin")
r_tax     = row; row = put(row, "Tax rate", 0.24, NF_PCT, "Effective / marginal tax")
r_da      = row; row = put(row, "D&A (% of revenue)", 0.03, NF_PCT, "Depreciation & amortization")
r_capex   = row; row = put(row, "CapEx (% of revenue)", 0.031, NF_PCT, "Capital expenditure")
r_nwc     = row; row = put(row, "Change in NWC (% of dRev)", 0.05, NF_PCT, "Working-cap investment on YoY rev increase")
r_intexp  = row; row = put(row, "Interest expense ($M)", 3800, NF_NUM, "TTM interest expense")
r_ni      = row; row = put(row, "Net income ($M)", 112010, NF_NUM, "TTM net income")
row += 1

# ---- BALANCE SHEET ----
row = grp("BALANCE SHEET", row)
r_debt    = row; row = put(row, "Total debt ($M)", 100000, NF_NUM, "Short + long term debt")
r_cash    = row; row = put(row, "Cash & equivalents ($M)", 132420, NF_NUM, "Cash + ST investments")
r_bve     = row; row = put(row, "Book value of equity ($M)", 73730, NF_NUM, "Total shareholders equity")
r_ta      = row; row = put(row, "Total assets ($M)", 359240, NF_NUM, "Total assets")
r_tca     = row; row = put(row, "Total current assets ($M)", 147957, NF_NUM, "Current assets")
r_tcl     = row; row = put(row, "Total current liabilities ($M)", 165631, NF_NUM, "Current liabilities")
r_re      = row; row = put(row, "Retained earnings ($M)", -14264, NF_NUM, "Retained earnings")
r_tl      = row; row = put(row, "Total liabilities ($M)", 285510, NF_NUM, "Total liabilities")
row += 1

# ---- GROWTH ----
row = grp("GROWTH", row)
r_g1 = row; row = put(row, "Revenue growth yr 1", 0.07, NF_PCT, "Explicit stage growth")
r_g2 = row; row = put(row, "Revenue growth yr 2", 0.06, NF_PCT, "")
r_g3 = row; row = put(row, "Revenue growth yr 3", 0.06, NF_PCT, "")
r_g4 = row; row = put(row, "Revenue growth yr 4", 0.05, NF_PCT, "")
r_g5 = row; row = put(row, "Revenue growth yr 5", 0.05, NF_PCT, "")
r_gt = row; row = put(row, "Terminal growth", 0.03, NF_PCT, "Perpetuity growth rate")
row += 1

# ---- WACC ----
row = grp("WACC INPUTS", row)
r_rf   = row; row = put(row, "Risk-free rate", 0.043, NF_PCT, "10y treasury")
r_erp  = row; row = put(row, "Equity risk premium", 0.05, NF_PCT, "Market risk premium")
r_kdpt = row; row = put(row, "Pre-tax cost of debt", 0.05, NF_PCT, "Yield on debt")
r_aaa  = row; row = put(row, "AAA corp bond yield", 0.045, NF_PCT, "For Graham revised formula")
row += 1

# ---- PEERS ----
row = grp("RELATIVE PEERS (MEDIANS)", row)
r_ppe   = row; row = put(row, "Peer P/E", 30, NF_MULT, "Trailing P/E median")
r_pfpe  = row; row = put(row, "Peer forward P/E", 27, NF_MULT, "Forward P/E median")
r_peve  = row; row = put(row, "Peer EV/EBITDA", 22, NF_MULT, "EV/EBITDA median")
r_pevs  = row; row = put(row, "Peer EV/Sales", 8, NF_MULT, "EV/Sales median")
r_ppb   = row; row = put(row, "Peer P/B", 45, NF_MULT, "Price/Book median")
r_ppfcf = row; row = put(row, "Peer P/FCF", 30, NF_MULT, "Price/FCF median")
row += 1

# ---- SCENARIO ----
row = grp("SCENARIO DELTAS", row)
r_exit_base = row; row = put(row, "Exit EV/EBITDA - BASE", 16, NF_MULT, "Base exit multiple")
r_exit_bear = row; row = put(row, "Exit EV/EBITDA - BEAR", 12, NF_MULT, "Bear exit multiple")
r_exit_bull = row; row = put(row, "Exit EV/EBITDA - BULL", 20, NF_MULT, "Bull exit multiple")
r_pb_bear = row; row = put(row, "Prob - BEAR", 0.25, NF_PCT, "Scenario probability")
r_pb_base = row; row = put(row, "Prob - BASE", 0.50, NF_PCT, "Scenario probability")
r_pb_bull = row; row = put(row, "Prob - BULL", 0.25, NF_PCT, "Scenario probability")
row += 1

# ---- MONTE CARLO ----
row = grp("MONTE CARLO SIGMAS", row)
r_sd_g   = row; row = put(row, "SD revenue growth (pts)", 0.025, NF_PCT, "Std dev applied to growth")
r_sd_w   = row; row = put(row, "SD WACC (pts)", 0.008, NF_PCT, "Std dev applied to WACC")
r_sd_m   = row; row = put(row, "SD EBIT margin (pts)", 0.020, NF_PCT, "Std dev applied to margin")
r_sd_gt  = row; row = put(row, "SD terminal growth (pts)", 0.005, NF_PCT, "Std dev applied to g_term")
r_loss   = row; row = put(row, "Kelly downside loss L", 0.30, NF_PCT, "Assumed loss if thesis wrong")
row += 1

# ---- PIOTROSKI prior-year booleans (editable, default mix summing to 7) ----
row = grp("PIOTROSKI INPUTS (1=pass)", row)
piotroski_first = row
piotroski_labels = [
    ("F1 ROA > 0", 1),
    ("F2 CFO > 0", 1),
    ("F3 dROA > 0 vs prior", 1),
    ("F4 CFO > Net income", 1),
    ("F5 Lower LT-debt ratio", 1),
    ("F6 Higher current ratio", 0),
    ("F7 No new shares issued", 1),
    ("F8 Higher gross margin", 1),
    ("F9 Higher asset turnover", 0),
]
for label, dv in piotroski_labels:
    lbl(A, f"B{row}", label, size=10, font_color=C_TEXT)
    inp(A, f"C{row}", dv, "0", "1 = criterion passes, 0 = fails")
    row += 1
piotroski_last = row - 1
row += 1

# ---- BENEISH 8 ratio inputs (default ~1.0) ----
row = grp("BENEISH M INPUTS (~1.0 = clean)", row)
beneish_first = row
beneish_labels = ["DSRI","GMI","AQI","SGI","DEPI","SGAI","TATA","LVGI"]
beneish_defaults = [1.0,1.0,1.0,1.0,1.0,1.0,0.0,1.0]
for label, dv in zip(beneish_labels, beneish_defaults):
    lbl(A, f"B{row}", label, size=10, font_color=C_TEXT)
    inp(A, f"C{row}", dv, "0.000", "Beneish ratio input")
    row += 1
beneish_last = row - 1
row += 1

# ---- VALUATION HISTORY (own multiple bands, ~10y) ----
row = grp("VALUATION HISTORY (OWN, ~10Y)", row)
r_pe_lo = row; row = put(row, "P/E - 10y low", 11.0, NF_MULT, "Lowest ~10y P/E")
r_pe_md = row; row = put(row, "P/E - 10y median", 24.0, NF_MULT, "Median ~10y P/E")
r_pe_hi = row; row = put(row, "P/E - 10y high", 39.0, NF_MULT, "Highest ~10y P/E")
r_ev_lo = row; row = put(row, "EV/EBITDA - 10y low", 9.0, NF_MULT, "Lowest ~10y EV/EBITDA")
r_ev_md = row; row = put(row, "EV/EBITDA - 10y median", 18.0, NF_MULT, "Median ~10y EV/EBITDA")
r_ev_hi = row; row = put(row, "EV/EBITDA - 10y high", 30.0, NF_MULT, "Highest ~10y EV/EBITDA")
row += 1

# ---- QUALITY & MOAT (drives required margin of safety) ----
row = grp("QUALITY & MOAT", row)
r_moat = row; row = put(row, "Moat rating (0-2)", 2, "0", "0 none, 1 narrow, 2 wide")
r_mos_wide = row; row = put(row, "Req. MoS - wide moat", 0.10, NF_PCT, "Discount needed before buying")
r_mos_solid = row; row = put(row, "Req. MoS - solid", 0.20, NF_PCT, "")
r_mos_avg = row; row = put(row, "Req. MoS - average", 0.35, NF_PCT, "")
r_mos_low = row; row = put(row, "Req. MoS - low quality", 0.50, NF_PCT, "")
row += 1

# ---- MARKET REGIME (whole-market context) ----
row = grp("MARKET REGIME (CONTEXT)", row)
r_cape = row; row = put(row, "Shiller CAPE (S&P 500)", 37.0, NF_MULT, "Cyclically-adjusted P/E")
r_cape_avg = row; row = put(row, "CAPE long-run average", 17.0, NF_MULT, "~150y mean is ~17")
r_mkt_ey = row; row = put(row, "Market fwd earnings yield", 0.045, NF_PCT, "S&P 500 forward E/P")
r_g_mkt = row; row = put(row, "Market LT earnings growth", 0.04, NF_PCT, "For implied ERP = E/P+g-Y")
r_ust10 = row; row = put(row, "10y Treasury yield", 0.043, NF_PCT, "Risk-free benchmark")
r_breadth = row; row = put(row, "% S&P above 200d MA", 0.55, NF_PCT, "Market breadth")
row += 1

# ---- RISK CALIBRATION (Damodaran: scale MoS by uncertainty & concentration) ----
row = grp("RISK CALIBRATION (DAMODARAN)", row)
r_uncert = row; row = put(row, "Estimate uncertainty (1-3)", 2, "0", "1 low (utility) - 3 high (uncertain)")
r_holdings = row; row = put(row, "Portfolio holdings (count)", 15, "0", "Fewer holdings -> higher required MoS")
row += 1

# ---- BLEND WEIGHTS (method mix; auto-renormalized) ----
row = grp("BLEND WEIGHTS (METHOD MIX)", row)
r_w_fcff = row; row = put(row, "Weight DCF-FCFF", 0.30, NF_PCT, "Edit to re-weight; auto-renormalized")
r_w_fcfe = row; row = put(row, "Weight DCF-FCFE", 0.15, NF_PCT, "")
r_w_rel = row; row = put(row, "Weight Relative", 0.25, NF_PCT, "")
r_w_epv = row; row = put(row, "Weight EPV", 0.20, NF_PCT, "")
r_w_gnum = row; row = put(row, "Weight Graham Number", 0.05, NF_PCT, "")
r_w_gform = row; row = put(row, "Weight Graham Formula", 0.05, NF_PCT, "")

# Build a reference dict to ASSUMPTIONS cells (absolute)
P = {}  # name -> "ASSUMPTIONS!$C$n"
def aref(r):
    return f"ASSUMPTIONS!$C${r}"
P.update(dict(
    price=aref(r_price), shares=aref(r_shares), hi52=aref(r_hi52), lo52=aref(r_lo52),
    swhi=aref(r_swhi), swlo=aref(r_swlo), ma50=aref(r_ma50), ma200=aref(r_ma200),
    rsi=aref(r_rsi), beta=aref(r_beta),
    rev=aref(r_rev), ebitm=aref(r_ebitm), tax=aref(r_tax), da=aref(r_da),
    capex=aref(r_capex), nwc=aref(r_nwc), intexp=aref(r_intexp), ni=aref(r_ni),
    debt=aref(r_debt), cash=aref(r_cash), bve=aref(r_bve), ta=aref(r_ta),
    tca=aref(r_tca), tcl=aref(r_tcl), re=aref(r_re), tl=aref(r_tl),
    g1=aref(r_g1), g2=aref(r_g2), g3=aref(r_g3), g4=aref(r_g4), g5=aref(r_g5),
    gt=aref(r_gt),
    rf=aref(r_rf), erp=aref(r_erp), kdpt=aref(r_kdpt), aaa=aref(r_aaa),
    ppe=aref(r_ppe), pfpe=aref(r_pfpe), peve=aref(r_peve), pevs=aref(r_pevs),
    ppb=aref(r_ppb), ppfcf=aref(r_ppfcf),
    exit_base=aref(r_exit_base), exit_bear=aref(r_exit_bear), exit_bull=aref(r_exit_bull),
    pb_bear=aref(r_pb_bear), pb_base=aref(r_pb_base), pb_bull=aref(r_pb_bull),
    sd_g=aref(r_sd_g), sd_w=aref(r_sd_w), sd_m=aref(r_sd_m), sd_gt=aref(r_sd_gt),
    loss=aref(r_loss),
))
P["piotroski_range"] = f"ASSUMPTIONS!$C${piotroski_first}:$C${piotroski_last}"
P["beneish"] = {name: aref(beneish_first + i) for i, name in enumerate(beneish_labels)}
P.update(dict(
    pe_lo=aref(r_pe_lo), pe_md=aref(r_pe_md), pe_hi=aref(r_pe_hi),
    ev_lo=aref(r_ev_lo), ev_md=aref(r_ev_md), ev_hi=aref(r_ev_hi),
    moat=aref(r_moat), mos_wide=aref(r_mos_wide), mos_solid=aref(r_mos_solid),
    mos_avg=aref(r_mos_avg), mos_low=aref(r_mos_low),
    cape=aref(r_cape), cape_avg=aref(r_cape_avg), mkt_ey=aref(r_mkt_ey),
    g_mkt=aref(r_g_mkt), ust10=aref(r_ust10), breadth=aref(r_breadth),
    uncert=aref(r_uncert), holdings=aref(r_holdings),
    w_fcff=aref(r_w_fcff), w_fcfe=aref(r_w_fcfe), w_rel=aref(r_w_rel),
    w_epv=aref(r_w_epv), w_gnum=aref(r_w_gnum), w_gform=aref(r_w_gform),
))

# ============================================================================
# 3) DCF  (WACC builder, FCFF projection, dual TV, FCFE, sensitivity)
# ============================================================================
D = ws_dcf
banner(D, "DISCOUNTED CASH FLOW", "2-stage FCFF with dual terminal value, FCFE mini-model, and WACC x g sensitivity.", "K")
no_grid_freeze(D, "A4")
paint_background(D, "K", 90)
setcols(D, {"A":3,"B":26,"C":13,"D":13,"E":13,"F":13,"G":13,"H":13,"I":13,"J":13,"K":13})

# ---- WACC BUILDER (rows 5-12) ----
section_label(D, "B5", "WACC BUILDER", fill=fill_raised)
for col in "CDEFGHIJK":
    D[f"{col}5"].fill = fill_raised
lbl(D, "B6", "Cost of equity Ke = Rf + Beta*ERP")
val(D, "C6", f"={P['rf']}+{P['beta']}*{P['erp']}", NF_PCT2, color=C_TEXT, bold=True)
lbl(D, "B7", "After-tax Kd (pre-tax = int/debt if avail)")
val(D, "C7", f"=IF(AND({P['intexp']}>0,{P['debt']}>0),{P['intexp']}/{P['debt']},{P['kdpt']})*(1-{P['tax']})", NF_PCT2)
lbl(D, "B8", "Equity value E = price*shares")
val(D, "C8", f"={P['price']}*{P['shares']}", NF_NUM)
lbl(D, "B9", "Debt value D = total debt")
val(D, "C9", f"={P['debt']}", NF_NUM)
lbl(D, "B10", "V = E + D")
val(D, "C10", "=C8+C9", NF_NUM)
lbl(D, "B11", "WACC = (E/V)*Ke + (D/V)*Kd_at")
val(D, "C11", "=(C8/C10)*C6+(C9/C10)*C7", NF_PCT2, color=C_GOLD_BR, bold=True)
WACC = "DCF!$C$11"
KE   = "DCF!$C$6"

lbl(D, "B12", "Net debt = total debt - cash")
val(D, "C12", f"={P['debt']}-{P['cash']}", NF_NUM)
NETDEBT = "DCF!$C$12"

# ---- FCFF PROJECTION (rows 14-26), years in cols D..H (1..5), base in C ----
section_label(D, "B14", "FCFF 2-STAGE PROJECTION ($M)", fill=fill_raised)
for col in "CDEFGHIJK":
    D[f"{col}14"].fill = fill_raised
hdr = ["", "Base (t0)", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
cols5 = ["C","D","E","F","G","H"]
style_cell(D, "B15", "Line item", font_color=C_MUTED, size=8, bold=True, fill=fill_raised)
for i, col in enumerate(cols5):
    style_cell(D, f"{col}15", hdr[i+1], font_color=C_MUTED, size=8, bold=True,
               fill=fill_raised, align="right")

gref = [P['g1'],P['g2'],P['g3'],P['g4'],P['g5']]
# Revenue row 16
lbl(D, "B16", "Revenue")
val(D, "C16", f"={P['rev']}", NF_NUM)
for i, col in enumerate(cols5[1:]):
    prev = cols5[i]  # cols5[i] is previous column
    val(D, f"{col}16", f"={prev}16*(1+{gref[i]})", NF_NUM)
# growth row 17
lbl(D, "B17", "  growth %")
for i, col in enumerate(cols5[1:]):
    val(D, f"{col}17", f"={gref[i]}", NF_PCT, color=C_MUTED)
# EBIT row 18
lbl(D, "B18", "EBIT")
for col in cols5[1:]:
    val(D, f"{col}18", f"={col}16*{P['ebitm']}", NF_NUM)
# NOPAT row 19
lbl(D, "B19", "NOPAT = EBIT*(1-tax)")
for col in cols5[1:]:
    val(D, f"{col}19", f"={col}18*(1-{P['tax']})", NF_NUM)
# D&A row 20
lbl(D, "B20", "D&A")
for col in cols5[1:]:
    val(D, f"{col}20", f"={col}16*{P['da']}", NF_NUM)
# CapEx row 21
lbl(D, "B21", "CapEx")
for col in cols5[1:]:
    val(D, f"{col}21", f"=-{col}16*{P['capex']}", NF_NUM, color=C_RED)
# dNWC row 22
lbl(D, "B22", "Change in NWC")
for i, col in enumerate(cols5[1:]):
    prev = cols5[i]
    val(D, f"{col}22", f"=-({col}16-{prev}16)*{P['nwc']}", NF_NUM, color=C_RED)
# FCFF row 23
lbl(D, "B23", "FCFF", size=10, font_color=C_TEXT)
for col in cols5[1:]:
    val(D, f"{col}23", f"={col}19+{col}20+{col}21+{col}22", NF_NUM,
        color=C_GREEN, bold=True)
# Discount factor row 24
lbl(D, "B24", "Discount factor")
for i, col in enumerate(cols5[1:]):
    t = i + 1
    val(D, f"{col}24", f"=1/(1+{WACC})^{t}", "0.0000", color=C_MUTED)
# PV of FCFF row 25
lbl(D, "B25", "PV of FCFF")
for col in cols5[1:]:
    val(D, f"{col}25", f"={col}23*{col}24", NF_NUM)
# Sum PV row 26
lbl(D, "B26", "Sum PV(FCFF)", size=10, font_color=C_TEXT)
val(D, "C26", "=SUM(D25:H25)", NF_NUM, color=C_GOLD_BR, bold=True)
PV_FCFF = "DCF!$C$26"
# base-year FCFF (FCFF_0) used by Owner earnings / P/FCF: compute on base
lbl(D, "B27", "FCFF base (t0)")
val(D, "C27",
    f"={P['rev']}*{P['ebitm']}*(1-{P['tax']})+{P['rev']}*{P['da']}-{P['rev']}*{P['capex']}",
    NF_NUM)
FCFF0 = "DCF!$C$27"

# ---- TERMINAL VALUE (rows 29-37) ----
section_label(D, "B29", "TERMINAL VALUE — DUAL METHOD", fill=fill_raised)
for col in "CDEFGHIJK":
    D[f"{col}29"].fill = fill_raised
lbl(D, "B30", "A. Gordon: FCFF5*(1+gt)/(WACC-gt)")
# guard WACC<=gt
val(D, "C30",
    f"=IF({WACC}<={P['gt']},0,H23*(1+{P['gt']})/({WACC}-{P['gt']}))",
    NF_NUM)
TV_A = "DCF!$C$30"
lbl(D, "B31", "EBITDA Year5 = EBIT5 + D&A5")
val(D, "C31", "=H18+H20", NF_NUM)
lbl(D, "B32", "B. Exit multiple: EBITDA5 * exit")
val(D, "C32", f"=C31*{P['exit_base']}", NF_NUM)
TV_B = "DCF!$C$32"
lbl(D, "B33", "Discount factor DF5")
val(D, "C33", f"=1/(1+{WACC})^5", "0.0000", color=C_MUTED)
DF5 = "DCF!$C$33"
lbl(D, "B34", "PV of TV (Gordon)")
val(D, "C34", "=C30*C33", NF_NUM)
lbl(D, "B35", "PV of TV (Exit)")
val(D, "C35", "=C32*C33", NF_NUM)
lbl(D, "B36", "Blended terminal (avg A,B)")
val(D, "C36", "=AVERAGE(C30,C32)", NF_NUM, color=C_TEXT)
lbl(D, "B37", "PV of blended TV")
val(D, "C37", "=C36*C33", NF_NUM, color=C_GOLD_BR, bold=True)
PV_TV = "DCF!$C$37"

# ---- EV -> EQUITY -> PER SHARE (rows 39-45) ----
section_label(D, "B39", "EV → EQUITY → PER-SHARE", fill=fill_raised)
for col in "CDEFGHIJK":
    D[f"{col}39"].fill = fill_raised
lbl(D, "B40", "Enterprise value = PV FCFF + PV TV")
val(D, "C40", "=C26+C37", NF_NUM, bold=True)
lbl(D, "B41", "less Net debt")
val(D, "C41", f"=-{NETDEBT}", NF_NUM, color=C_RED)
lbl(D, "B42", "Equity value")
val(D, "C42", "=C40+C41", NF_NUM, bold=True)
lbl(D, "B43", "DCF value per share (FCFF)")
val(D, "C43", f"=C42/{P['shares']}", NF_USD, color=C_GOLD_BR, bold=True, size=12)
DCF_FCFF_PS = "DCF!$C$43"
lbl(D, "B44", "Upside vs price")
val(D, "C44", f"=C43/{P['price']}-1", NF_PCT,
    color=C_GREEN)

# ---- FCFE MINI-MODEL (rows 46-56) ----
section_label(D, "B46", "FCFE MINI-MODEL (discount at Ke)", fill=fill_raised)
for col in "CDEFGHIJK":
    D[f"{col}46"].fill = fill_raised
style_cell(D, "B47", "Line item", font_color=C_MUTED, size=8, bold=True, fill=fill_raised)
for i, col in enumerate(cols5):
    style_cell(D, f"{col}47", hdr[i+1], font_color=C_MUTED, size=8, bold=True,
               fill=fill_raised, align="right")
# Net income proj: scale base NI by revenue growth path (proxy)
lbl(D, "B48", "Net income (grown w/ rev)")
val(D, "C48", f"={P['ni']}", NF_NUM)
for i, col in enumerate(cols5[1:]):
    prev = cols5[i]
    val(D, f"{col}48", f"={prev}48*(1+{gref[i]})", NF_NUM)
lbl(D, "B49", "+ D&A")
for col in cols5[1:]:
    val(D, f"{col}49", f"={col}20", NF_NUM)
lbl(D, "B50", "- CapEx")
for col in cols5[1:]:
    val(D, f"{col}50", f"={col}21", NF_NUM, color=C_RED)
lbl(D, "B51", "- Change in NWC")
for col in cols5[1:]:
    val(D, f"{col}51", f"={col}22", NF_NUM, color=C_RED)
lbl(D, "B52", "+ Net borrowing (0 default)")
for col in cols5[1:]:
    val(D, f"{col}52", "=0", NF_NUM, color=C_MUTED)
lbl(D, "B53", "FCFE", size=10, font_color=C_TEXT)
for col in cols5[1:]:
    val(D, f"{col}53", f"={col}48+{col}49+{col}50+{col}51+{col}52", NF_NUM,
        color=C_GREEN, bold=True)
lbl(D, "B54", "PV(FCFE) @ Ke")
for i, col in enumerate(cols5[1:]):
    t = i + 1
    val(D, f"{col}54", f"={col}53/(1+{KE})^{t}", NF_NUM)
lbl(D, "B55", "Terminal FCFE (Gordon @ Ke)")
val(D, "C55",
    f"=IF({KE}<={P['gt']},NA(),H53*(1+{P['gt']})/({KE}-{P['gt']})/(1+{KE})^5)",
    NF_NUM)
lbl(D, "B56", "FCFE equity value")
val(D, "C56", "=SUM(D54:H54)+C55", NF_NUM, bold=True)
lbl(D, "B57", "FCFE value per share")
val(D, "C57", f"=C56/{P['shares']}", NF_USD, color=C_GOLD_BR, bold=True, size=12)
DCF_FCFE_PS = "DCF!$C$57"

# ---- SENSITIVITY TABLE (WACC rows x terminal-growth cols) ----
# Build a computed grid (not a Data Table) so it always recalcs.
# WACC from base-1.5% to +1.5% step 0.5% (7 rows); gt from -1% to +1% step 0.5% (5 cols)
section_label(D, "B60", "SENSITIVITY: DCF $/SH  (WACC ↓  x  g_term →)", fill=fill_raised)
for col in "CDEFGHIJK":
    D[f"{col}60"].fill = fill_raised
# header corner
style_cell(D, "B61", "WACC \\ g", font_color=C_MUTED, size=8, bold=True, fill=fill_raised)
gt_offsets = [-0.01, -0.005, 0.0, 0.005, 0.01]
sens_cols = ["C","D","E","F","G"]
for j, off in enumerate(gt_offsets):
    style_cell(D, f"{sens_cols[j]}61", f"={P['gt']}+({off})", numfmt=NF_PCT,
               font_color=C_MUTED, size=8, bold=True, fill=fill_raised, align="right")
wacc_offsets = [-0.015,-0.01,-0.005,0.0,0.005,0.01,0.015]
sens_first_row = 62
for i, woff in enumerate(wacc_offsets):
    r = sens_first_row + i
    # row header = WACC value
    style_cell(D, f"B{r}", f"={WACC}+({woff})", numfmt=NF_PCT2, font_color=C_MUTED,
               size=9, bold=True, fill=fill_raised, align="right")
    for j, goff in enumerate(gt_offsets):
        col = sens_cols[j]
        w = f"({WACC}+({woff}))"
        g = f"({P['gt']}+({goff}))"
        # reduced-form DCF per share recomputed at (w,g):
        # PV FCFF uses actual FCFF rows discounted at w; TV via Gordon at (w,g).
        pv = (f"(D23/(1+{w})^1 + E23/(1+{w})^2 + F23/(1+{w})^3 + "
              f"G23/(1+{w})^4 + H23/(1+{w})^5)")
        tv = f"IF({w}<={g},NA(),H23*(1+{g})/({w}-{g})/(1+{w})^5)"
        formula = f"=(({pv})+({tv})-{NETDEBT})/{P['shares']}"
        val(D, f"{col}{r}", formula, NF_USD, color=C_TEXT)
sens_last_row = sens_first_row + len(wacc_offsets) - 1
# color scale green-amber-red across the grid
D.conditional_formatting.add(
    f"C{sens_first_row}:G{sens_last_row}",
    ColorScaleRule(start_type="min", start_color=C_RED,
                   mid_type="percentile", mid_value=50, mid_color=C_AMBER,
                   end_type="max", end_color=C_GREEN))

# ============================================================================
# 4) RELATIVE  (current vs peer multiples; implied price per method)
# ============================================================================
Rl = ws_rel
banner(Rl, "RELATIVE VALUATION", "Implied price from peer multiples vs current multiples; PEG.", "H")
no_grid_freeze(Rl, "A4")
paint_background(Rl, "H", 60)
setcols(Rl, {"A":3,"B":26,"C":14,"D":14,"E":14,"F":14,"G":3,"H":3})

# Helper EBITDA / EPS / forward EPS / BVPS as references
EPS    = f"({P['ni']}/{P['shares']})"
EBITDA = f"({P['rev']}*{P['ebitm']}+{P['rev']}*{P['da']})"
FWDEPS = f"({EPS}*(1+{P['g1']}))"
BVPS   = f"({P['bve']}/{P['shares']})"
FCFPS  = f"({FCFF0}/{P['shares']})"

section_label(Rl, "B5", "MULTIPLES — CURRENT vs PEER → IMPLIED PRICE", fill=fill_raised)
for col in "CDEF":
    Rl[f"{col}5"].fill = fill_raised
heads = ["Method","Current","Peer median","Implied $/sh","Upside %"]
hcols = ["B","C","D","E","F"]
for i, col in enumerate(hcols):
    style_cell(Rl, f"{col}6", heads[i], font_color=C_MUTED, size=8, bold=True,
               fill=fill_raised, align=("left" if i==0 else "right"))

def rel_row(r, name, current_f, peer_ref, implied_f, cfmt=NF_MULT):
    lbl(Rl, f"B{r}", name, size=10, font_color=C_TEXT)
    val(Rl, f"C{r}", current_f, cfmt)
    val(Rl, f"D{r}", f"={peer_ref}", NF_MULT, color=C_MUTED)
    val(Rl, f"E{r}", implied_f, NF_USD, color=C_GOLD_BR, bold=True)
    val(Rl, f"F{r}", f"=E{r}/{P['price']}-1", NF_PCT)

# P/E
rel_row(7, "P/E", f"={P['price']}/{EPS}", P['ppe'], f"={P['ppe']}*{EPS}")
# Forward P/E
rel_row(8, "Forward P/E", f"={P['price']}/{FWDEPS}", P['pfpe'], f"={P['pfpe']}*{FWDEPS}")
# EV/EBITDA
lbl(Rl, "B9", "EV/EBITDA", size=10, font_color=C_TEXT)
val(Rl, "C9", f"=({P['price']}*{P['shares']}+{P['debt']}-{P['cash']})/{EBITDA}", NF_MULT)
val(Rl, "D9", f"={P['peve']}", NF_MULT, color=C_MUTED)
val(Rl, "E9", f"=({P['peve']}*{EBITDA}-({P['debt']}-{P['cash']}))/{P['shares']}", NF_USD, color=C_GOLD_BR, bold=True)
val(Rl, "F9", f"=E9/{P['price']}-1", NF_PCT)
# EV/Sales
lbl(Rl, "B10", "EV/Sales", size=10, font_color=C_TEXT)
val(Rl, "C10", f"=({P['price']}*{P['shares']}+{P['debt']}-{P['cash']})/{P['rev']}", NF_MULT)
val(Rl, "D10", f"={P['pevs']}", NF_MULT, color=C_MUTED)
val(Rl, "E10", f"=({P['pevs']}*{P['rev']}-({P['debt']}-{P['cash']}))/{P['shares']}", NF_USD, color=C_GOLD_BR, bold=True)
val(Rl, "F10", f"=E10/{P['price']}-1", NF_PCT)
# P/B
rel_row(11, "P/B", f"={P['price']}/{BVPS}", P['ppb'], f"={P['ppb']}*{BVPS}")
# P/FCF
rel_row(12, "P/FCF", f"={P['price']}/{FCFPS}", P['ppfcf'], f"={P['ppfcf']}*{FCFPS}")

# PEG
section_label(Rl, "B14", "PEG", fill=fill_raised)
for col in "CDEF":
    Rl[f"{col}14"].fill = fill_raised
lbl(Rl, "B15", "PEG = (P/E)/(g1*100)")
val(Rl, "C15", f"=({P['price']}/{EPS})/({P['g1']}*100)", "0.00", color=C_TEXT, bold=True)

# Avg of P/E-implied and EV/EBITDA-implied (used by dashboard)
lbl(Rl, "B17", "Relative blended (P/E & EV/EBITDA)", size=10, font_color=C_TEXT)
val(Rl, "C17", "=AVERAGE(E7,E9)", NF_USD, color=C_GOLD_BR, bold=True)
REL_BLEND = "RELATIVE!$C$17"

# small bar of implied prices
chart = BarChart()
chart.type = "col"
chart.title = "Implied price by method ($/sh)"
chart.height = 6
chart.width = 14
data = Reference(Rl, min_col=5, min_row=7, max_row=12)
cats = Reference(Rl, min_col=2, min_row=7, max_row=12)
chart.add_data(data, titles_from_data=False)
chart.set_categories(cats)
chart.legend = None
# Terminal aesthetic: category axis along the BOTTOM, value axis at LEFT, no gridlines.
chart.x_axis.axPos = "b"
chart.y_axis.axPos = "l"
chart.x_axis.delete = False
chart.y_axis.delete = False
chart.x_axis.majorGridlines = None
chart.y_axis.majorGridlines = None
Rl.add_chart(chart, "B20")

# ============================================================================
# 5) VALUE_HEURISTICS  (Graham, EPV, Owner earnings, Reverse DCF)
# ============================================================================
V = ws_val
banner(V, "VALUE HEURISTICS", "Graham Number & Formula, EPV, Owner Earnings, and the Reverse DCF solver.", "I")
no_grid_freeze(V, "A4")
paint_background(V, "I", 70)
setcols(V, {"A":3,"B":30,"C":14,"D":14,"E":14,"F":14,"G":14,"H":14,"I":3})

section_label(V, "B5", "HEURISTIC FAIR VALUES ($/SH)", fill=fill_raised)
for col in "CD":
    V[f"{col}5"].fill = fill_raised
# Graham Number
lbl(V, "B6", "Graham Number = sqrt(22.5*EPS*BVPS)")
val(V, "C6",
    f"=IF(AND({EPS}>0,{BVPS}>0),SQRT(22.5*{EPS}*{BVPS}),NA())",
    NF_USD, color=C_GOLD_BR, bold=True)
GRAHAM_NUM = "VALUE_HEURISTICS!$C$6"
# Graham Revised Formula
lbl(V, "B7", "Graham Formula = EPS*(8.5+2*g)*4.4/Y")
val(V, "C7",
    f"={EPS}*(8.5+2*MIN({P['g1']}*100,15))*4.4/({P['aaa']}*100)",
    NF_USD, color=C_GOLD_BR, bold=True)
GRAHAM_FORM = "VALUE_HEURISTICS!$C$7"
# EPV
lbl(V, "B8", "EPV = NOPAT/WACC - NetDebt (per sh)")
val(V, "C8",
    f"=IF({WACC}<=0,NA(),(({P['rev']}*{P['ebitm']}*(1-{P['tax']}))/{WACC}-{NETDEBT})/{P['shares']})",
    NF_USD, color=C_GOLD_BR, bold=True)
EPV_PS = "VALUE_HEURISTICS!$C$8"
# Owner earnings
section_label(V, "B10", "OWNER EARNINGS (BUFFETT)", fill=fill_raised)
for col in "CD":
    V[f"{col}10"].fill = fill_raised
lbl(V, "B11", "Owner earnings = NI + D&A - maint capex")
# maintenance capex = D&A
val(V, "C11", f"={P['ni']}+{P['rev']}*{P['da']}-0.70*{P['rev']}*{P['capex']}", NF_NUM, color=C_TEXT, bold=True)
lbl(V, "B12", "  (maint capex ~ 70% of total capex)")
lbl(V, "B13", "Owner earnings yield = OE / mkt cap")
val(V, "C13", f"=C11/({P['price']}*{P['shares']})", NF_PCT2, color=C_GREEN, bold=True)

# ---- REVERSE DCF SOLVER TABLE ----
section_label(V, "B16", "REVERSE DCF — GROWTH THE MARKET IS PRICING", fill=fill_raised)
for col in "CDEFGH":
    V[f"{col}16"].fill = fill_raised
lbl(V, "B17", "Solve g* so DCF $/sh = current price (g_term fixed).")
# Build a vertical solver table: g from -10% to 30% step 2% in column B, value col C
style_cell(V, "B18", "g (5yr)", font_color=C_MUTED, size=8, bold=True, fill=fill_raised)
style_cell(V, "C18", "DCF $/sh at g", font_color=C_MUTED, size=8, bold=True, fill=fill_raised, align="right")
style_cell(V, "D18", "value - price", font_color=C_MUTED, size=8, bold=True, fill=fill_raised, align="right")
g_steps = [(-0.10 + 0.02*i) for i in range(21)]  # -10%..30%
solver_first = 19
# Full dual-terminal DCF at FLAT growth g (closed form) — mirrors the HTML reverse-DCF
# (5-yr FCFF, blended Gordon+Exit terminal) so the two files reconcile.
_mgn = P['ebitm']; _tax = P['tax']; _da = P['da']; _capex = P['capex']; _nwc = P['nwc']
w = WACC; gt = P['gt']; _exitb = P['exit_base']
for i, g in enumerate(g_steps):
    r = solver_first + i
    style_cell(V, f"B{r}", round(g, 4), font_color=C_TEXT, size=9, mono=True,
               fill=fill_panel, align="right", numfmt=NF_PCT, border=border_hair)
    gref = f"B{r}"
    A   = f"({_mgn}*(1-{_tax})+{_da}-{_capex})"                  # FCFF margin ex-NWC
    Bc  = f"({A}-{_nwc}*{gref}/(1+{gref}))"                       # incl. NWC drag
    q   = f"((1+{gref})/(1+{w}))"
    sumpv = f"{P['rev']}*{Bc}*{q}*(1-{q}^5)/(1-{q})"             # PV of 5 growing FCFF
    rev5  = f"{P['rev']}*(1+{gref})^5"
    fcff5 = f"({rev5}*{Bc})"
    ebd5  = f"({rev5}*({_mgn}+{_da}))"
    gordon = f"IF({w}<={gt},0,{fcff5}*(1+{gt})/({w}-{gt}))"
    exitt  = f"{ebd5}*{_exitb}"
    pvterm = f"((({gordon})+({exitt}))/2)/(1+{w})^5"             # blended terminal, discounted
    val(V, f"C{r}", f"=(({sumpv})+({pvterm})-{NETDEBT})/{P['shares']}", NF_USD, color=C_TEXT)
    val(V, f"D{r}", f"=C{r}-{P['price']}", NF_USD, color=C_TEXT)
solver_last = solver_first + len(g_steps) - 1
# Implied g* via LOCAL interpolation at the zero-crossing of (value - price). Column C is
# monotone increasing in g, so MATCH(price, C, 1) brackets the root; then linear-interpolate.
section_label(V, "F18", "IMPLIED vs BASE", fill=fill_raised)
V["F18"].fill = fill_raised
lbl(V, "F19", "Implied g* (market)")
_Cr = f"C{solver_first}:C{solver_last}"; _Br = f"B{solver_first}:B{solver_last}"
_m  = f"MATCH({P['price']},{_Cr},1)"
_interp = (f"INDEX({_Br},{_m})+({P['price']}-INDEX({_Cr},{_m}))*"
           f"(INDEX({_Br},{_m}+1)-INDEX({_Br},{_m}))/(INDEX({_Cr},{_m}+1)-INDEX({_Cr},{_m}))")
val(V, "G19",
    f"=IFERROR({_interp},IF({P['price']}<INDEX({_Cr},1),B{solver_first},B{solver_last}))",
    NF_PCT, color=C_GOLD_BR, bold=True)
IMPLIED_G = "VALUE_HEURISTICS!$G$19"
lbl(V, "F20", "Base-case g (yr1)")
val(V, "G20", f"={P['g1']}", NF_PCT, color=C_TEXT, bold=True)
lbl(V, "F21", "Market - base (gap)")
val(V, "G21", f"=G19-G20", NF_PCT, color=C_AMBER, bold=True)
lbl(V, "F22", "Read")
style_cell(V, "G22",
           '=IF(G19<G20,"Market pricing LESS growth than base → cheap","Market pricing MORE growth → rich")',
           font_color=C_MUTED, size=8, italic=True, fill=fill_panel, align="right", wrap=True)

# ============================================================================
# 6) QUALITY  (ROIC vs WACC, Piotroski, Altman, Beneish)
# ============================================================================
Q = ws_qual
banner(Q, "QUALITY & SAFETY", "ROIC vs WACC, Piotroski F-Score, Altman Z, Beneish M.", "H")
no_grid_freeze(Q, "A4")
paint_background(Q, "H", 75)
setcols(Q, {"A":3,"B":34,"C":13,"D":13,"E":13,"F":13,"G":13,"H":3})

# ---- ROIC ----
section_label(Q, "B5", "ROIC vs WACC", fill=fill_raised)
for col in "CD":
    Q[f"{col}5"].fill = fill_raised
NOPAT0 = f"({P['rev']}*{P['ebitm']}*(1-{P['tax']}))"
lbl(Q, "B6", "NOPAT (current)")
val(Q, "C6", f"={NOPAT0}", NF_NUM)
lbl(Q, "B7", "Invested capital = Debt+Equity-Cash")
val(Q, "C7", f"={P['debt']}+{P['bve']}-{P['cash']}", NF_NUM)
lbl(Q, "B8", "ROIC = NOPAT / invested capital")
val(Q, "C8", "=IF(C7=0,NA(),C6/C7)", NF_PCT2, color=C_GOLD_BR, bold=True)
ROIC = "QUALITY!$C$8"
lbl(Q, "B9", "WACC")
val(Q, "C9", f"={WACC}", NF_PCT2)
lbl(Q, "B10", "Spread = ROIC - WACC")
val(Q, "C10", "=C8-C9", NF_PCT2, color=C_GREEN, bold=True)
lbl(Q, "B11", "Verdict")
style_cell(Q, "C11", '=IF(C8>C9,"VALUE-CREATING","VALUE-DESTROYING")',
           font_color=C_TEXT, size=10, bold=True, mono=True, fill=fill_panel,
           align="right", border=border_hair)
Q.conditional_formatting.add("C11",
    FormulaRule(formula=['$C$8>$C$9'], fill=fill_green, font=Font(name=F_MONO, bold=True, color=C_BG)))
Q.conditional_formatting.add("C11",
    FormulaRule(formula=['$C$8<=$C$9'], fill=fill_red, font=Font(name=F_MONO, bold=True, color=C_TEXT)))

# ---- PIOTROSKI ----
section_label(Q, "B13", "PIOTROSKI F-SCORE (0-9)", fill=fill_raised)
for col in "CD":
    Q[f"{col}13"].fill = fill_raised
style_cell(Q, "B14", "Criterion", font_color=C_MUTED, size=8, bold=True, fill=fill_raised)
style_cell(Q, "C14", "Pass(1/0)", font_color=C_MUTED, size=8, bold=True, fill=fill_raised, align="right")
crit_names = [
    "F1  ROA > 0","F2  CFO > 0","F3  dROA > 0 vs prior","F4  CFO > Net income",
    "F5  Lower LT-debt ratio","F6  Higher current ratio","F7  No new shares issued",
    "F8  Higher gross margin","F9  Higher asset turnover"]
piox_first = 15
for i, name in enumerate(crit_names):
    r = piox_first + i
    lbl(Q, f"B{r}", name, size=9, font_color=C_TEXT)
    val(Q, f"C{r}", f"=ASSUMPTIONS!$C${piotroski_first+i}", "0", color=C_TEXT)
    Q.conditional_formatting.add(f"C{r}",
        CellIsRule(operator="equal", formula=["1"], fill=fill_green,
                   font=Font(name=F_MONO, bold=True, color=C_BG)))
    Q.conditional_formatting.add(f"C{r}",
        CellIsRule(operator="equal", formula=["0"], fill=fill_red,
                   font=Font(name=F_MONO, bold=True, color=C_TEXT)))
piox_last = piox_first + len(crit_names) - 1
piox_total_row = piox_last + 1
lbl(Q, f"B{piox_total_row}", "F-Score total", size=10, font_color=C_TEXT)
val(Q, f"C{piox_total_row}", f"=SUM(C{piox_first}:C{piox_last})", "0",
    color=C_GOLD_BR, bold=True, size=12)
PIOTROSKI = f"QUALITY!$C${piox_total_row}"
Q.conditional_formatting.add(f"C{piox_total_row}",
    CellIsRule(operator="greaterThanOrEqual", formula=["7"], fill=fill_green,
               font=Font(name=F_MONO, bold=True, color=C_BG)))
Q.conditional_formatting.add(f"C{piox_total_row}",
    CellIsRule(operator="between", formula=["4","6"], fill=fill_amber,
               font=Font(name=F_MONO, bold=True, color=C_BG)))
Q.conditional_formatting.add(f"C{piox_total_row}",
    CellIsRule(operator="lessThan", formula=["4"], fill=fill_red,
               font=Font(name=F_MONO, bold=True, color=C_TEXT)))

# ---- ALTMAN Z ----
altman_start = piox_total_row + 2
section_label(Q, f"B{altman_start}", "ALTMAN Z-SCORE (MANUFACTURING)", fill=fill_raised)
for col in "CD":
    Q[f"{col}{altman_start}"].fill = fill_raised
ar = altman_start + 1
lbl(Q, f"B{ar}", "A = (CA-CL)/TA")
val(Q, f"C{ar}", f"=({P['tca']}-{P['tcl']})/{P['ta']}", "0.000")
lbl(Q, f"B{ar+1}", "B = RetainedEarnings/TA")
val(Q, f"C{ar+1}", f"={P['re']}/{P['ta']}", "0.000")
lbl(Q, f"B{ar+2}", "C = EBIT/TA")
val(Q, f"C{ar+2}", f"=({P['rev']}*{P['ebitm']})/{P['ta']}", "0.000")
lbl(Q, f"B{ar+3}", "D = MktCapEquity/TotalLiab")
val(Q, f"C{ar+3}", f"=({P['price']}*{P['shares']})/{P['tl']}", "0.000")
lbl(Q, f"B{ar+4}", "E = Revenue/TA")
val(Q, f"C{ar+4}", f"={P['rev']}/{P['ta']}", "0.000")
lbl(Q, f"B{ar+5}", "Z = 1.2A+1.4B+3.3C+0.6D+1.0E", size=10, font_color=C_TEXT)
val(Q, f"C{ar+5}", f"=1.2*C{ar}+1.4*C{ar+1}+3.3*C{ar+2}+0.6*C{ar+3}+1.0*C{ar+4}",
    "0.00", color=C_GOLD_BR, bold=True, size=12)
ALTMAN = f"QUALITY!$C${ar+5}"
lbl(Q, f"B{ar+6}", "Zone")
style_cell(Q, f"C{ar+6}",
           f'=IF(C{ar+5}>2.99,"SAFE",IF(C{ar+5}>=1.81,"GREY","DISTRESS"))',
           font_color=C_TEXT, size=10, bold=True, mono=True, fill=fill_panel,
           align="right", border=border_hair)
ALTMAN_ZONE = f"QUALITY!$C${ar+6}"
Q.conditional_formatting.add(f"C{ar+6}",
    FormulaRule(formula=[f'$C${ar+5}>2.99'], fill=fill_green, font=Font(name=F_MONO, bold=True, color=C_BG)))
Q.conditional_formatting.add(f"C{ar+6}",
    FormulaRule(formula=[f'AND($C${ar+5}>=1.81,$C${ar+5}<=2.99)'], fill=fill_amber, font=Font(name=F_MONO, bold=True, color=C_BG)))
Q.conditional_formatting.add(f"C{ar+6}",
    FormulaRule(formula=[f'$C${ar+5}<1.81'], fill=fill_red, font=Font(name=F_MONO, bold=True, color=C_TEXT)))

# ---- BENEISH M ----
ben_start = ar + 8
section_label(Q, f"B{ben_start}", "BENEISH M-SCORE", fill=fill_raised)
for col in "CD":
    Q[f"{col}{ben_start}"].fill = fill_raised
bcoef = [("DSRI",0.920),("GMI",0.528),("AQI",0.404),("SGI",0.892),
         ("DEPI",0.115),("SGAI",-0.172),("TATA",4.679),("LVGI",-0.327)]
br0 = ben_start + 1
style_cell(Q, f"B{br0}", "Ratio", font_color=C_MUTED, size=8, bold=True, fill=fill_raised)
style_cell(Q, f"C{br0}", "Input", font_color=C_MUTED, size=8, bold=True, fill=fill_raised, align="right")
style_cell(Q, f"D{br0}", "Coef", font_color=C_MUTED, size=8, bold=True, fill=fill_raised, align="right")
beneish_terms = []
for i, (name, coef) in enumerate(bcoef):
    r = br0 + 1 + i
    lbl(Q, f"B{r}", name, size=9, font_color=C_TEXT)
    val(Q, f"C{r}", f"={P['beneish'][name]}", "0.000", color=C_TEXT)
    val(Q, f"D{r}", coef, "0.000", color=C_MUTED)
    beneish_terms.append(f"{coef}*C{r}")
ben_m_row = br0 + 1 + len(bcoef)
lbl(Q, f"B{ben_m_row}", "M = -4.84 + sum(coef*ratio)", size=10, font_color=C_TEXT)
val(Q, f"C{ben_m_row}", "=-4.84+" + "+".join(beneish_terms),
    "0.000", color=C_GOLD_BR, bold=True, size=12)
BENEISH = f"QUALITY!$C${ben_m_row}"
lbl(Q, f"B{ben_m_row+1}", "Flag (M > -1.78 = manipulation)")
style_cell(Q, f"C{ben_m_row+1}",
           f'=IF(C{ben_m_row}>-1.78,"FLAG ⚠","CLEAN")',
           font_color=C_TEXT, size=10, bold=True, mono=True, fill=fill_panel,
           align="right", border=border_hair)
BENEISH_FLAG = f"QUALITY!$C${ben_m_row+1}"
Q.conditional_formatting.add(f"C{ben_m_row+1}",
    FormulaRule(formula=[f'$C${ben_m_row}>-1.78'], fill=fill_red, font=Font(name=F_MONO, bold=True, color=C_TEXT)))
Q.conditional_formatting.add(f"C{ben_m_row+1}",
    FormulaRule(formula=[f'$C${ben_m_row}<=-1.78'], fill=fill_green, font=Font(name=F_MONO, bold=True, color=C_BG)))

# ============================================================================
# 7) TIMING  (Fib ladder, MAs, RSI, 52w pos, MoS bands, DCA planner)
# ============================================================================
T = ws_time
banner(T, "TIMING & ENTRY", "Fibonacci ladder, moving averages, RSI, 52-week position, MoS bands, DCA planner.", "I")
no_grid_freeze(T, "A4")
paint_background(T, "I", 80)
setcols(T, {"A":3,"B":26,"C":14,"D":14,"E":16,"F":14,"G":14,"H":14,"I":3})

# Blended fair value reference (defined on dashboard) -- we reference it
BLENDED_FV = "DASHBOARD!$C$8"   # set later on dashboard

# ---- FIBONACCI LADDER ----
section_label(T, "B5", "FIBONACCI RETRACEMENT LADDER", fill=fill_raised)
for col in "CDE":
    T[f"{col}5"].fill = fill_raised
lbl(T, "B6", "Swing high H")
val(T, "C6", f"={P['swhi']}", NF_USD)
lbl(T, "B7", "Swing low L")
val(T, "C7", f"={P['swlo']}", NF_USD)
lbl(T, "B8", "Range R = H - L")
val(T, "C8", "=C6-C7", NF_USD)
style_cell(T, "B10", "Level", font_color=C_MUTED, size=8, bold=True, fill=fill_raised)
style_cell(T, "C10", "Price", font_color=C_MUTED, size=8, bold=True, fill=fill_raised, align="right")
style_cell(T, "D10", "vs price", font_color=C_MUTED, size=8, bold=True, fill=fill_raised, align="right")
fib_levels = [("0.000 (H)",0.0),("0.236",0.236),("0.382",0.382),("0.500",0.5),
              ("0.618",0.618),("0.786",0.786),("1.000 (L)",1.0),
              ("1.272 ext",1.272),("1.618 ext",1.618)]
fib_first = 11
for i,(name,k) in enumerate(fib_levels):
    r = fib_first + i
    lbl(T, f"B{r}", name, size=9, font_color=C_TEXT)
    if k <= 1.0:
        val(T, f"C{r}", f"=C6-C8*{k}", NF_USD, color=C_TEXT)
    else:
        val(T, f"C{r}", f"=C7-C8*({k}-1)", NF_USD, color=C_BLUE)
    val(T, f"D{r}", f"=C{r}-{P['price']}", NF_USD, color=C_MUTED)
fib_last = fib_first + len(fib_levels) - 1
# nearest support below price
nss = fib_last + 2
lbl(T, f"B{nss}", "Nearest Fib support < price", size=10, font_color=C_TEXT)
# largest fib level that is <= price
val(T, f"C{nss}",
    f"=MAX(IF(C{fib_first}:C{fib_last}<={P['price']},C{fib_first}:C{fib_last}))",
    NF_USD, color=C_GREEN, bold=True)
NEAR_FIB = f"TIMING!$C${nss}"
# Non-array, no-CSE: nearest Fib level at or below price via MAXIFS (Excel 2019+/365).
# Stored with the mandatory _xlfn. prefix so real Excel + LibreOffice both resolve it.
T[f"C{nss}"].value = (
    f"=_xlfn.MAXIFS(C{fib_first}:C{fib_last},C{fib_first}:C{fib_last},\"<=\"&{P['price']})")
# highlight nearest support row
T.conditional_formatting.add(f"C{fib_first}:C{fib_last}",
    FormulaRule(formula=[f'C{fib_first}=$C${nss}'], fill=fill_green,
                font=Font(name=F_MONO, bold=True, color=C_BG)))

# ---- MOVING AVERAGES ----
ma_s = nss + 2
section_label(T, f"B{ma_s}", "MOVING AVERAGES & TREND", fill=fill_raised)
for col in "CD":
    T[f"{col}{ma_s}"].fill = fill_raised
lbl(T, f"B{ma_s+1}", "Price")
val(T, f"C{ma_s+1}", f"={P['price']}", NF_USD)
lbl(T, f"B{ma_s+2}", "50-day MA")
val(T, f"C{ma_s+2}", f"={P['ma50']}", NF_USD)
lbl(T, f"B{ma_s+3}", "200-day MA")
val(T, f"C{ma_s+3}", f"={P['ma200']}", NF_USD)
lbl(T, f"B{ma_s+4}", "Cross signal")
style_cell(T, f"C{ma_s+4}",
           f'=IF({P["ma50"]}>{P["ma200"]},"GOLDEN CROSS","DEATH CROSS")',
           font_color=C_TEXT, size=10, bold=True, mono=True, fill=fill_panel,
           align="right", border=border_hair)
T.conditional_formatting.add(f"C{ma_s+4}",
    FormulaRule(formula=[f'{P["ma50"]}>{P["ma200"]}'], fill=fill_green, font=Font(name=F_MONO, bold=True, color=C_BG)))
T.conditional_formatting.add(f"C{ma_s+4}",
    FormulaRule(formula=[f'{P["ma50"]}<={P["ma200"]}'], fill=fill_red, font=Font(name=F_MONO, bold=True, color=C_TEXT)))
lbl(T, f"B{ma_s+5}", "Trend vs MAs")
style_cell(T, f"C{ma_s+5}",
           f'=IF(AND({P["price"]}>{P["ma50"]},{P["price"]}>{P["ma200"]}),"UPTREND",IF(AND({P["price"]}<{P["ma50"]},{P["price"]}<{P["ma200"]}),"DOWNTREND","MIXED"))',
           font_color=C_TEXT, size=10, bold=True, mono=True, fill=fill_panel, align="right", border=border_hair)

# ---- RSI + 52w position (to the right, col F/G) ----
section_label(T, "F5", "RSI & 52-WEEK POSITION", fill=fill_raised)
for col in "GH":
    T[f"{col}5"].fill = fill_raised
lbl(T, "F6", "RSI(14)")
val(T, "G6", f"={P['rsi']}", NF_NUM1, color=C_GOLD_BR, bold=True)
lbl(T, "F7", "RSI state")
style_cell(T, "G7",
           f'=IF({P["rsi"]}<30,"OVERSOLD",IF({P["rsi"]}>70,"OVERBOUGHT","NEUTRAL"))',
           font_color=C_TEXT, size=10, bold=True, mono=True, fill=fill_panel, align="right", border=border_hair)
T.conditional_formatting.add("G7",
    FormulaRule(formula=[f'{P["rsi"]}<30'], fill=fill_green, font=Font(name=F_MONO, bold=True, color=C_BG)))
T.conditional_formatting.add("G7",
    FormulaRule(formula=[f'{P["rsi"]}>70'], fill=fill_red, font=Font(name=F_MONO, bold=True, color=C_TEXT)))
lbl(T, "F8", "52w range position")
val(T, "G8", f"=({P['price']}-{P['lo52']})/({P['hi52']}-{P['lo52']})", NF_PCT,
    color=C_TEXT, bold=True)
lbl(T, "F9", "  0%=low  100%=high")

# ---- MARGIN OF SAFETY BANDS ----
mos_s = ma_s + 7
section_label(T, f"B{mos_s}", "MARGIN-OF-SAFETY BANDS", fill=fill_raised)
for col in "CD":
    T[f"{col}{mos_s}"].fill = fill_raised
style_cell(T, f"B{mos_s+1}", "MoS", font_color=C_MUTED, size=8, bold=True, fill=fill_raised)
style_cell(T, f"C{mos_s+1}", "Buy price", font_color=C_MUTED, size=8, bold=True, fill=fill_raised, align="right")
for i, m in enumerate([0.15,0.30,0.50]):
    r = mos_s + 2 + i
    lbl(T, f"B{r}", f"{int(m*100)}% MoS", size=9, font_color=C_TEXT)
    val(T, f"C{r}", f"={BLENDED_FV}*(1-{m})", NF_USD, color=C_GREEN, bold=True)

# ---- DCA TRANCHE PLANNER ----
dca_s = mos_s + 7
section_label(T, f"F{mos_s}", "DCA 4-TRANCHE PLANNER", fill=fill_raised)
for col in "GH":
    T[f"{col}{mos_s}"].fill = fill_raised
style_cell(T, f"F{mos_s+1}", "Tranche", font_color=C_MUTED, size=8, bold=True, fill=fill_raised)
style_cell(T, f"G{mos_s+1}", "Buy price", font_color=C_MUTED, size=8, bold=True, fill=fill_raised, align="right")
style_cell(T, f"H{mos_s+1}", "Weight", font_color=C_MUTED, size=8, bold=True, fill=fill_raised, align="right")
# Tranche prices: current price, Fib 0.382, 0.5, 0.618 (or FV*(1-x) if lower)
fib382 = f"(C6-C8*0.382)"
fib500 = f"(C6-C8*0.5)"
fib618 = f"(C6-C8*0.618)"
tranches = [
    ("T1 @ market", f"={P['price']}", 0.40),
    ("T2 @ Fib .382", f"=MIN({fib382},{BLENDED_FV}*(1-0.10))", 0.30),
    ("T3 @ Fib .500", f"=MIN({fib500},{BLENDED_FV}*(1-0.20))", 0.20),
    ("T4 @ Fib .618", f"=MIN({fib618},{BLENDED_FV}*(1-0.30))", 0.10),
]
dca_first = mos_s + 2
for i,(name,pf,w) in enumerate(tranches):
    r = dca_first + i
    lbl(T, f"F{r}", name, size=9, font_color=C_TEXT)
    val(T, f"G{r}", pf, NF_USD, color=C_TEXT)
    val(T, f"H{r}", w, NF_PCT, color=C_MUTED)
dca_last = dca_first + 3
blend_row = dca_last + 1
lbl(T, f"F{blend_row}", "Blended avg cost", size=10, font_color=C_TEXT)
val(T, f"G{blend_row}", f"=SUMPRODUCT(G{dca_first}:G{dca_last},H{dca_first}:H{dca_last})",
    NF_USD, color=C_GOLD_BR, bold=True)
lbl(T, f"F{blend_row+1}", "Discount to FV")
val(T, f"G{blend_row+1}", f"=1-G{blend_row}/{BLENDED_FV}", NF_PCT, color=C_GREEN, bold=True)

# ============================================================================
# 8) SCENARIO_MC  (scenarios, Monte Carlo, Kelly)
# ============================================================================
M = ws_mc
banner(M, "SCENARIO & MONTE CARLO", "Bull/Base/Bear DCF, 2000-trial Monte Carlo, percentiles, and Kelly sizing.", "S")
no_grid_freeze(M, "A4")
paint_background(M, "S", 60)
setcols(M, {"A":3,"B":26,"C":14,"D":14,"E":14,"F":14,"G":4,
            "H":13,"I":13,"J":13,"K":13,"L":13,"M":4,"N":14,"O":14,"P":4,
            "Q":30,"R":13,"S":13})

# Helper: scenario reduced-form per-share value.
# Use constant-growth proxy: avg explicit growth +- delta, EBIT margin +- delta, WACC +- delta, exit multiple.
# Build a clean scenario table with explicit recompute using reduced form:
#   FCFF0 -> grow at g for 5y, discount at w, TV via exit multiple on EBITDA5.
section_label(M, "B5", "SCENARIO FAIR VALUES ($/SH)", fill=fill_raised)
for col in "CDEF":
    M[f"{col}5"].fill = fill_raised
style_cell(M, "B6", "Driver", font_color=C_MUTED, size=8, bold=True, fill=fill_raised)
style_cell(M, "C6", "BEAR", font_color=C_RED, size=8, bold=True, fill=fill_raised, align="right")
style_cell(M, "D6", "BASE", font_color=C_GOLD_BR, size=8, bold=True, fill=fill_raised, align="right")
style_cell(M, "E6", "BULL", font_color=C_GREEN, size=8, bold=True, fill=fill_raised, align="right")

# base average growth
AVG_G = f"AVERAGE({P['g1']},{P['g2']},{P['g3']},{P['g4']},{P['g5']})"
# row 7: growth
lbl(M, "B7", "Avg growth g")
val(M, "C7", f"={AVG_G}-0.03", NF_PCT, color=C_TEXT)
val(M, "D7", f"={AVG_G}", NF_PCT, color=C_TEXT)
val(M, "E7", f"={AVG_G}+0.03", NF_PCT, color=C_TEXT)
# row 8: EBIT margin
lbl(M, "B8", "EBIT margin")
val(M, "C8", f"={P['ebitm']}-0.03", NF_PCT, color=C_TEXT)
val(M, "D8", f"={P['ebitm']}", NF_PCT, color=C_TEXT)
val(M, "E8", f"={P['ebitm']}+0.02", NF_PCT, color=C_TEXT)
# row 9: WACC
lbl(M, "B9", "WACC")
val(M, "C9", f"={WACC}+0.01", NF_PCT2, color=C_TEXT)
val(M, "D9", f"={WACC}", NF_PCT2, color=C_TEXT)
val(M, "E9", f"={WACC}-0.005", NF_PCT2, color=C_TEXT)
# row 10: exit multiple
lbl(M, "B10", "Exit EV/EBITDA")
val(M, "C10", f"={P['exit_bear']}", NF_MULT, color=C_TEXT)
val(M, "D10", f"={P['exit_base']}", NF_MULT, color=C_TEXT)
val(M, "E10", f"={P['exit_bull']}", NF_MULT, color=C_TEXT)

# ---- FULL 5-YEAR FCFF PER SCENARIO ----
# Mirrors the main DCF engine exactly (tapering per-year growth, full FCFF incl.
# dNWC, blended Gordon+Exit terminal) so the BASE column ties to DCF!C43 to the
# cent and the prob-weighted value reconciles with the HTML cockpit.
# Drivers already on rows 7-10: row8 = scenario EBIT margin, row9 = scenario
# WACC, row10 = scenario exit multiple. Growth delta vs base per year:
scen_cols = ["C", "D", "E"]
dg = {"C": "-0.03", "D": "", "E": "+0.03"}            # growth shift per year
gref_s = [P['g1'], P['g2'], P['g3'], P['g4'], P['g5']]
yr_lbl = ["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]

# Revenue by year (label row 12, years rows 13-17)
lbl(M, "B12", "REVENUE ($M)", size=8, font_color=C_MUTED)
for t in range(1, 6):
    r = 12 + t  # 13..17
    lbl(M, f"B{r}", f"  {yr_lbl[t-1]}", size=9, font_color=C_MUTED)
    for col in scen_cols:
        prev = P['rev'] if t == 1 else f"{col}{r-1}"
        val(M, f"{col}{r}", f"={prev}*(1+{gref_s[t-1]}{dg[col]})", NF_NUM, color=C_MUTED)

# FCFF by year (label row 19, years rows 20-24): NOPAT + D&A - CapEx - dNWC
lbl(M, "B19", "FCFF ($M)", size=8, font_color=C_MUTED)
for t in range(1, 6):
    rr = 19 + t       # 20..24
    rev_r = 12 + t    # matching revenue row
    lbl(M, f"B{rr}", f"  {yr_lbl[t-1]}", size=9, font_color=C_MUTED)
    for col in scen_cols:
        revc = f"{col}{rev_r}"
        prevrev = P['rev'] if t == 1 else f"{col}{rev_r-1}"
        f = (f"={revc}*{col}8*(1-{P['tax']})+{revc}*{P['da']}"
             f"-{revc}*{P['capex']}-({revc}-{prevrev})*{P['nwc']}")
        val(M, f"{col}{rr}", f, NF_NUM, color=C_GREEN if t == 5 else C_TEXT)

# Valuation bridge (rows 26-34), discounting at each scenario's own WACC (row 9)
lbl(M, "B26", "Sum PV(FCFF)")
for col in scen_cols:
    pv = "+".join([f"{col}{19+t}/(1+{col}9)^{t}" for t in range(1, 6)])
    val(M, f"{col}26", f"={pv}", NF_NUM, color=C_TEXT)
lbl(M, "B27", "EBITDA Year 5")
for col in scen_cols:
    val(M, f"{col}27", f"={col}17*({col}8+{P['da']})", NF_NUM, color=C_MUTED)
lbl(M, "B28", "Terminal - Gordon")
for col in scen_cols:
    val(M, f"{col}28",
        f"=IF({col}9<={P['gt']},0,{col}24*(1+{P['gt']})/({col}9-{P['gt']}))",
        NF_NUM, color=C_MUTED)
lbl(M, "B29", "Terminal - Exit mult")
for col in scen_cols:
    val(M, f"{col}29", f"={col}27*{col}10", NF_NUM, color=C_MUTED)
lbl(M, "B30", "Blended terminal (avg)")
for col in scen_cols:
    val(M, f"{col}30", f"=AVERAGE({col}28,{col}29)", NF_NUM, color=C_TEXT)
lbl(M, "B31", "PV of terminal")
for col in scen_cols:
    val(M, f"{col}31", f"={col}30/(1+{col}9)^5", NF_NUM, color=C_TEXT)
lbl(M, "B32", "Enterprise value")
for col in scen_cols:
    val(M, f"{col}32", f"={col}26+{col}31", NF_NUM, color=C_TEXT)
lbl(M, "B33", "Equity value")
for col in scen_cols:
    val(M, f"{col}33", f"={col}32-{NETDEBT}", NF_NUM, color=C_TEXT)
lbl(M, "B34", "Fair value $/sh", size=10, font_color=C_TEXT)
for col, color in (("C", C_RED), ("D", C_GOLD_BR), ("E", C_GREEN)):
    val(M, f"{col}34", f"={col}33/{P['shares']}", NF_USD, color=color, bold=True, size=12)
SC_BEAR = "SCENARIO_MC!$C$34"
SC_BASE = "SCENARIO_MC!$D$34"
SC_BULL = "SCENARIO_MC!$E$34"
# Probabilities (row 35) and probability-weighted fair value (row 36)
lbl(M, "B35", "Probability")
val(M, "C35", f"={P['pb_bear']}", NF_PCT, color=C_MUTED)
val(M, "D35", f"={P['pb_base']}", NF_PCT, color=C_MUTED)
val(M, "E35", f"={P['pb_bull']}", NF_PCT, color=C_MUTED)
lbl(M, "B36", "Prob-weighted fair value", size=10, font_color=C_TEXT)
val(M, "C36", "=C34*C35+D34*D35+E34*E35",
    NF_USD, color=C_GOLD_BR, bold=True, size=12)
SC_PW = "SCENARIO_MC!$C$36"

# ---- MONTE CARLO ENGINE (helper area cols H..L, 2000 trials) ----
section_label(M, "H5", "MONTE CARLO ENGINE — 2000 TRIALS", fill=fill_raised)
for col in "IJKL":
    M[f"{col}5"].fill = fill_raised
mc_heads = ["g","WACC","margin","g_term","FV $/sh"]
mc_cols = ["H","I","J","K","L"]
for i,col in enumerate(mc_cols):
    style_cell(M, f"{col}6", mc_heads[i], font_color=C_MUTED, size=8, bold=True,
               fill=fill_raised, align="right")
mc_first = 7
mc_n = 2000
mc_last = mc_first + mc_n - 1
# mu references
mu_g   = AVG_G
mu_w   = WACC
mu_m   = P['ebitm']
mu_gt  = P['gt']
for i in range(mc_n):
    r = mc_first + i
    # g  (NORM.INV is a post-2007 function: store WITH the _xlfn. prefix)
    M[f"H{r}"] = f"=_xlfn.NORM.INV(RAND(),{mu_g},{P['sd_g']})"
    # WACC floored above g_term+0.5% to avoid divide error
    M[f"I{r}"] = f"=MAX(_xlfn.NORM.INV(RAND(),{mu_w},{P['sd_w']}),K{r}+0.005)"
    # margin
    M[f"J{r}"] = f"=_xlfn.NORM.INV(RAND(),{mu_m},{P['sd_m']})"
    # terminal growth
    M[f"K{r}"] = f"=_xlfn.NORM.INV(RAND(),{mu_gt},{P['sd_gt']})"
    # reduced-form FV: FCFF0(margin) growing at g 5y discounted at w + TV Gordon at gt
    fcff0 = f"({P['rev']}*J{r}*(1-{P['tax']})+{P['rev']}*{P['da']}-{P['rev']}*{P['capex']})"
    fcff1 = f"{fcff0}*(1+H{r})"
    pv = f"{fcff1}*(1-((1+H{r})/(1+I{r}))^5)/(I{r}-H{r})"
    fcff5 = f"{fcff0}*(1+H{r})^5"
    tv = f"{fcff5}*(1+K{r})/(I{r}-K{r})/(1+I{r})^5"
    M[f"L{r}"] = f"=(({pv})+({tv})-{NETDEBT})/{P['shares']}"
    # style only a few rows fully to keep file light; format the FV col
    for col in mc_cols:
        cc = M[f"{col}{r}"]
        cc.font = Font(name=F_MONO, size=8, color=C_MUTED)
        cc.number_format = (NF_USD if col=="L" else NF_PCT)
        cc.alignment = Alignment(horizontal="right")

# ---- PERCENTILE SUMMARY (cols N/O) ----
section_label(M, "N5", "MC SUMMARY", fill=fill_raised)
M["O5"].fill = fill_raised
fvrange = f"L{mc_first}:L{mc_last}"
summ = [
    ("P5",  f"=_xlfn.PERCENTILE.INC({fvrange},0.05)"),
    ("P25", f"=_xlfn.PERCENTILE.INC({fvrange},0.25)"),
    ("P50 (median)", f"=_xlfn.PERCENTILE.INC({fvrange},0.50)"),
    ("P75", f"=_xlfn.PERCENTILE.INC({fvrange},0.75)"),
    ("P95", f"=_xlfn.PERCENTILE.INC({fvrange},0.95)"),
    ("Mean", f"=AVERAGE({fvrange})"),
    ("P(FV > price)", f"=IFERROR(COUNTIF({fvrange},\">\"&{P['price']})/{mc_n},NA())"),
]
sr0 = 6
for i,(name,f) in enumerate(summ):
    r = sr0 + i
    lbl(M, f"N{r}", name, size=9, font_color=C_TEXT)
    fmt = NF_PCT if name.startswith("P(") else NF_USD
    val(M, f"O{r}",
        f, fmt, color=(C_GREEN if name.startswith("P(") else C_GOLD_BR), bold=True)
MC_P50 = f"SCENARIO_MC!$O${sr0+2}"
MC_PUNDER = f"SCENARIO_MC!$O${sr0+6}"

# ---- HISTOGRAM via FREQUENCY ----
hist_s = sr0 + 8
section_label(M, f"N{hist_s}", "HISTOGRAM BINS", fill=fill_raised)
M[f"O{hist_s}"].fill = fill_raised
style_cell(M, f"N{hist_s+1}", "Bin upper", font_color=C_MUTED, size=8, bold=True, fill=fill_raised, align="right")
style_cell(M, f"O{hist_s+1}", "Count", font_color=C_MUTED, size=8, bold=True, fill=fill_raised, align="right")
# 12 bins from P5 to P95
bin_first = hist_s + 2
nbins = 12
for i in range(nbins):
    r = bin_first + i
    frac = (i + 1) / nbins
    val(M, f"N{r}",
        f"=MIN({fvrange})+({frac})*(MAX({fvrange})-MIN({fvrange}))",
        NF_USD, color=C_MUTED)
bin_last = bin_first + nbins - 1
# FREQUENCY array over bins -> counts in O
freq_formula = f"=FREQUENCY({fvrange},N{bin_first}:N{bin_last})"
for i in range(nbins):
    r = bin_first + i
    M[f"O{r}"] = freq_formula  # will become array; we set as CSE below
    cc = M[f"O{r}"]
    cc.font = Font(name=F_MONO, size=9, color=C_TEXT)
    cc.number_format = NF_NUM
    cc.alignment = Alignment(horizontal="right")
# Mark as a single CSE array spilling N rows: openpyxl writes plain; Excel re-enters.
# Provide a robust per-bin COUNTIFS fallback so the histogram populates even without CSE:
for i in range(nbins):
    r = bin_first + i
    if i == 0:
        M[f"O{r}"] = f"=COUNTIF({fvrange},\"<=\"&N{r})"
    else:
        M[f"O{r}"] = f"=COUNTIFS({fvrange},\">\"&N{r-1},{fvrange},\"<=\"&N{r})"

# histogram bar chart
hist_chart = BarChart()
hist_chart.type = "col"
hist_chart.title = "Monte Carlo: fair value distribution"
hist_chart.height = 7
hist_chart.width = 16
hdata = Reference(M, min_col=15, min_row=bin_first, max_row=bin_last)  # col O
hcats = Reference(M, min_col=14, min_row=bin_first, max_row=bin_last)  # col N
hist_chart.add_data(hdata, titles_from_data=False)
hist_chart.set_categories(hcats)
hist_chart.legend = None
# Terminal aesthetic: category axis along the BOTTOM, value axis at LEFT, no gridlines.
hist_chart.x_axis.axPos = "b"
hist_chart.y_axis.axPos = "l"
hist_chart.x_axis.delete = False
hist_chart.y_axis.delete = False
hist_chart.x_axis.majorGridlines = None
hist_chart.y_axis.majorGridlines = None
M.add_chart(hist_chart, "B22")

# ---- KELLY SIZING ----
# IMPORTANT: the Monte Carlo trial block occupies H7:L2006. Keep the Kelly panel
# in columns Q/R/S (label/value/raw) so it never overwrites a trial input cell —
# an overlap on H..L would inject text into the MC formulas and yield #VALUE!.
kelly_s = 22
kL, kV, kR = "Q", "R", "S"   # label, value, raw-uncapped columns
section_label(M, f"{kL}{kelly_s}", "KELLY POSITION SIZING", fill=fill_raised)
for col in (kV, kR):
    M[f"{col}{kelly_s}"].fill = fill_raised
# blended MoS comes from dashboard
BLENDED_MOS = "DASHBOARD!$C$10"   # margin of safety on dashboard
lbl(M, f"{kL}{kelly_s+1}", "p = clamp(0.5+MoS,0.10,0.90)")
val(M, f"{kV}{kelly_s+1}", f"=MIN(0.90,MAX(0.10,0.5+{BLENDED_MOS}))", NF_PCT2, color=C_TEXT, bold=True)
lbl(M, f"{kL}{kelly_s+2}", "b = FV/price - 1 (upside)")
val(M, f"{kV}{kelly_s+2}", f"={BLENDED_FV}/{P['price']}-1", NF_PCT2, color=C_TEXT)
lbl(M, f"{kL}{kelly_s+3}", "L = downside loss")
val(M, f"{kV}{kelly_s+3}", f"={P['loss']}", NF_PCT2, color=C_TEXT)
lbl(M, f"{kL}{kelly_s+4}", "f* = (p*b-(1-p)*L)/(b*L)")
val(M, f"{kV}{kelly_s+4}",
    f"=IF({kV}{kelly_s+2}*{kV}{kelly_s+3}=0,0,({kV}{kelly_s+1}*{kV}{kelly_s+2}-(1-{kV}{kelly_s+1})*{kV}{kelly_s+3})/({kV}{kelly_s+2}*{kV}{kelly_s+3}))",
    NF_PCT, color=C_GOLD_BR, bold=True)
KELLY_FULL = f"SCENARIO_MC!${kV}${kelly_s+4}"
# Show BOTH the raw (uncapped) Kelly fraction and the capped recommendation,
# so a reader sees WHY Full and Half both land on the 25% ceiling (not a copy-paste bug).
style_cell(M, f"{kR}{kelly_s+4}", "raw f*", font_color=C_MUTED, size=8, bold=True,
           fill=fill_raised, align="right", border=border_hair)
lbl(M, f"{kL}{kelly_s+5}", "Half-Kelly = f*/2  ->  capped 25%")
val(M, f"{kV}{kelly_s+5}", f"=MIN(0.25,MAX(0,{kV}{kelly_s+4}/2))", NF_PCT, color=C_GREEN, bold=True)
val(M, f"{kR}{kelly_s+5}", f"=MAX(0,{kV}{kelly_s+4}/2)", NF_PCT, color=C_MUTED)
KELLY_HALF = f"SCENARIO_MC!${kV}${kelly_s+5}"
lbl(M, f"{kL}{kelly_s+6}", "Full Kelly = f*  ->  capped 25%")
val(M, f"{kV}{kelly_s+6}", f"=MIN(0.25,MAX(0,{kV}{kelly_s+4}))", NF_PCT, color=C_TEXT, bold=True)
val(M, f"{kR}{kelly_s+6}", f"=MAX(0,{kV}{kelly_s+4})", NF_PCT, color=C_MUTED)

# ============================================================================
# CONTEXT & REGIME  (built before DASHBOARD so the verdict can read the MoS bar)
#   A) valuation vs the stock's OWN ~10y multiple range
#   B) quality tier -> required margin of safety
#   C) market regime (CAPE, ERP, breadth)
#   D) effective required MoS = quality bar + regime adjustment
# ============================================================================
X = ws_context
banner(X, "CONTEXT & REGIME", "Valuation vs own history, quality-tiered margin of safety, and market regime.", "F")
no_grid_freeze(X, "A4")
paint_background(X, "F", 48)
setcols(X, {"A": 3, "B": 38, "C": 18, "D": 3, "E": 50, "F": 3})

def xtxt(cell, formula, color=C_TEXT, bold=False, size=10):
    val(X, cell, formula, None, color=color, bold=bold, size=size)
    X[cell].number_format = "General"

# ---- A) VALUATION vs OWN HISTORY ----
section_label(X, "B5", "VALUATION vs OWN HISTORY (~10Y)", fill=fill_raised)
for col in "CDE":
    X[f"{col}5"].fill = fill_raised
lbl(X, "B6", "Current P/E"); val(X, "C6", "=RELATIVE!$C$7", NF_MULT, color=C_TEXT)
lbl(X, "B7", "  P/E position in 10y range")
val(X, "C7", f'=MIN(1,MAX(0,(C6-{P["pe_lo"]})/MAX(0.01,{P["pe_hi"]}-{P["pe_lo"]})))', NF_PCT, color=C_MUTED)
lbl(X, "B8", "Current EV/EBITDA"); val(X, "C8", "=RELATIVE!$C$9", NF_MULT, color=C_TEXT)
lbl(X, "B9", "  EV/EBITDA position in 10y range")
val(X, "C9", f'=MIN(1,MAX(0,(C8-{P["ev_lo"]})/MAX(0.01,{P["ev_hi"]}-{P["ev_lo"]})))', NF_PCT, color=C_MUTED)
lbl(X, "B10", "Avg position (0=cheap, 1=rich)", size=10, font_color=C_TEXT)
val(X, "C10", "=AVERAGE(C7,C9)", NF_PCT, color=C_GOLD_BR, bold=True)
lbl(X, "B11", "Avg discount to own median (P/E & EV/EBITDA)")
val(X, "C11", f'=((C6/{P["pe_md"]}-1)+(C8/{P["ev_md"]}-1))/2', NF_PCT, color=C_TEXT)
lbl(X, "B12", "Valuation vs own history", size=10, font_color=C_TEXT)
xtxt("C12", '=IF(C10<0.25,"CHEAP",IF(C10<0.6,"FAIR",IF(C10<0.85,"RICH","EXTREME")))', color=C_GOLD_BR, bold=True)
VALHIST_POS = "CONTEXT!$C$10"; VALHIST_LABEL = "CONTEXT!$C$12"

# ---- B) QUALITY TIER -> REQUIRED MARGIN OF SAFETY ----
section_label(X, "B14", "QUALITY TIER -> REQUIRED MARGIN OF SAFETY", fill=fill_raised)
for col in "CDE":
    X[f"{col}14"].fill = fill_raised
lbl(X, "B15", "Moat rating (0-2)"); val(X, "C15", f"={P['moat']}", "0", color=C_TEXT)
lbl(X, "B16", "ROIC - WACC spread"); val(X, "C16", f"={ROIC}-{WACC}", NF_PCT2, color=C_GREEN)
lbl(X, "B17", "Piotroski F-Score"); val(X, "C17", f"={PIOTROSKI}", "0", color=C_TEXT)
lbl(X, "B18", "Quality tier", size=10, font_color=C_TEXT)
xtxt("C18",
     f'=IF(AND({P["moat"]}>=2,C16>=0.05,C17>=6),"WIDE-MOAT",'
     f'IF(AND({P["moat"]}>=1,C16>0,C17>=5),"SOLID",'
     f'IF(OR(C16>0,C17>=4),"AVERAGE","LOW")))', color=C_GOLD_BR, bold=True)
lbl(X, "B19", "Required margin of safety (base)", size=10, font_color=C_TEXT)
val(X, "C19",
    f'=IF(C18="WIDE-MOAT",{P["mos_wide"]},IF(C18="SOLID",{P["mos_solid"]},'
    f'IF(C18="AVERAGE",{P["mos_avg"]},{P["mos_low"]})))', NF_PCT, color=C_GOLD_BR, bold=True)
QUAL_TIER = "CONTEXT!$C$18"; REQ_MOS = "CONTEXT!$C$19"

# ---- C) MARKET REGIME ----
# ERP shown two ways: the academically-flawed Fed model (E/P - Y) labelled as a
# rough proxy, plus a more defensible implied ERP that adds expected growth.
section_label(X, "B21", "MARKET REGIME (CONTEXT)", fill=fill_raised)
for col in "CDE":
    X[f"{col}21"].fill = fill_raised
lbl(X, "B22", "Shiller CAPE (S&P 500)"); val(X, "C22", f"={P['cape']}", NF_MULT, color=C_TEXT)
lbl(X, "B23", "CAPE vs long-run avg"); val(X, "C23", f"=C22/{P['cape_avg']}-1", NF_PCT, color=C_TEXT)
lbl(X, "B24", "CAPE-implied 10y real return"); val(X, "C24", "=1/C22", NF_PCT, color=C_MUTED)
lbl(X, "B25", "Fed-model ERP (E/P - 10y, rough)"); val(X, "C25", f"={P['mkt_ey']}-{P['ust10']}", NF_PCT, color=C_MUTED)
lbl(X, "B26", "Implied ERP (E/P + g - 10y)"); val(X, "C26", f"={P['mkt_ey']}+{P['g_mkt']}-{P['ust10']}", NF_PCT, color=C_TEXT)
lbl(X, "B27", "% S&P above 200d MA"); val(X, "C27", f"={P['breadth']}", NF_PCT, color=C_TEXT)
lbl(X, "B28", "Market regime", size=10, font_color=C_TEXT)
xtxt("C28",
     f'=IF(OR(C22>={P["cape_avg"]}*1.6,C26<=0.02),"EXPENSIVE - raise bar",'
     f'IF(AND(C22<={P["cape_avg"]}*1.1,C26>=0.05),"CHEAP - be greedy","NEUTRAL"))', color=C_GOLD_BR, bold=True)
lbl(X, "B29", "Regime MoS adjustment")
val(X, "C29",
    f'=IF(OR(C22>={P["cape_avg"]}*1.6,C26<=0.02),0.05,IF(AND(C22<={P["cape_avg"]}*1.1,C26>=0.05),-0.05,0))',
    NF_PCT, color=C_MUTED)
REGIME_LABEL = "CONTEXT!$C$28"; REGIME_ADJ = "CONTEXT!$C$29"

# ---- D) RISK CALIBRATION (Damodaran: scale MoS by uncertainty + concentration) ----
section_label(X, "B31", "RISK CALIBRATION (DAMODARAN)", fill=fill_raised)
for col in "CDE":
    X[f"{col}31"].fill = fill_raised
lbl(X, "B32", "Estimate uncertainty (1 low - 3 high)"); val(X, "C32", f"={P['uncert']}", "0", color=C_TEXT)
lbl(X, "B33", "Uncertainty MoS add (+7.5%/level)")
val(X, "C33", f"=(MAX(1,MIN(3,{P['uncert']}))-1)*0.075", NF_PCT, color=C_MUTED)
lbl(X, "B34", "Portfolio holdings (count)"); val(X, "C34", f"={P['holdings']}", "0", color=C_TEXT)
lbl(X, "B35", "Concentration MoS add")
val(X, "C35", f'=IF({P["holdings"]}<=6,0.05,IF({P["holdings"]}<=12,0.025,0))', NF_PCT, color=C_MUTED)
UNCERT_ADJ = "CONTEXT!$C$33"; CONC_ADJ = "CONTEXT!$C$35"

# ---- E) EFFECTIVE REQUIRED MARGIN OF SAFETY (quality + regime + uncertainty + concentration) ----
section_label(X, "B37", "QUALITY-ADJUSTED ENTRY BAR", fill=fill_raised)
for col in "CDE":
    X[f"{col}37"].fill = fill_raised
lbl(X, "B38", "Required MoS (quality tier)"); val(X, "C38", "=C19", NF_PCT, color=C_TEXT)
lbl(X, "B39", "+ regime / uncertainty / concentration")
val(X, "C39", "=C29+C33+C35", NF_PCT, color=C_MUTED)
lbl(X, "B40", "Required MoS (effective)", size=10, font_color=C_TEXT)
val(X, "C40", "=MAX(0,C19+C29+C33+C35)", NF_PCT, color=C_GOLD_BR, bold=True, size=12)
lbl(X, "B41", "Your margin of safety"); val(X, "C41", "=DASHBOARD!$C$10", NF_PCT, color=C_TEXT)
lbl(X, "B42", "Meets bar to buy?", size=10, font_color=C_TEXT)
xtxt("C42", '=IF(C41>=C40,"YES - entry OK","NO - wait")', color=C_GREEN, bold=True)
REQ_MOS_EFF = "CONTEXT!$C$40"

# ============================================================================
# 1) DASHBOARD  (hero sheet, references everything)  -- build last
# ============================================================================
Dh = ws_dash
banner(Dh, "APEX TECHNOLOGIES (SAMPLE)", "Equity Valuation Cockpit — verdict, fair value, entry zone, quality and timing at a glance.", "K")
no_grid_freeze(Dh, "A4")
paint_background(Dh, "K", 60)
setcols(Dh, {"A":3,"B":24,"C":16,"D":16,"E":4,"F":22,"G":16,"H":16,"I":4,"J":3,"K":3})

# ---- BLENDED FAIR VALUE (compute here so other sheets can reference C8) ----
# Weights: FCFF 30, FCFE 15, Relative(avg P/E & EV/EBITDA) 25, EPV 20, GrahamNum 5, GrahamForm 5
# Renormalize over valid (non-NA) methods using IFERROR -> 0 weight.
section_label(Dh, "B5", "VALUATION VERDICT", fill=fill_raised)
for col in "CD":
    Dh[f"{col}5"].fill = fill_raised

# helper rows for blended fair value (place in a small block at right, cols F-H rows 5+ used for verdict)
# Compute weighted blended FV with NA-safe renormalization in C8.
# Build numerator and denominator using IFERROR.
methods = [
    (DCF_FCFF_PS, P['w_fcff']),
    (DCF_FCFE_PS, P['w_fcfe']),
    (REL_BLEND,   P['w_rel']),
    (EPV_PS,      P['w_epv']),
    (GRAHAM_NUM,  P['w_gnum']),
    (GRAHAM_FORM, P['w_gform']),
]
num = "+".join([f"IF(IFERROR({m},-1)>0,{m}*{w},0)" for m, w in methods])
den = "+".join([f"IF(IFERROR({m},-1)>0,{w},0)" for m, w in methods])

lbl(Dh, "B6", "Current price", size=10, font_color=C_TEXT)
val(Dh, "C6", f"={P['price']}", NF_USD, color=C_TEXT, bold=True, size=14)
lbl(Dh, "B7", "Blended fair value")
# C8 holds blended FV
lbl(Dh, "B8", "Blended fair value", size=10, font_color=C_TEXT)
val(Dh, "C8", f"=IF(({den})=0,NA(),({num})/({den}))", NF_USD, color=C_GOLD_BR, bold=True, size=14)
BLENDED = "DASHBOARD!$C$8"
lbl(Dh, "B9", "Upside vs price", size=10, font_color=C_TEXT)
val(Dh, "C9", f"=C8/C6-1", NF_PCT, color=C_GREEN, bold=True, size=12)
lbl(Dh, "B10", "Margin of safety", size=10, font_color=C_TEXT)
val(Dh, "C10", "=(C8-C6)/C8", NF_PCT, color=C_GREEN, bold=True, size=12)

# Quality gates for verdict
lbl(Dh, "B11", "Altman zone")
val(Dh, "C11", f"={ALTMAN_ZONE}", None, color=C_TEXT)
Dh["C11"].number_format = "General"
lbl(Dh, "B12", "Beneish flag")
val(Dh, "C12", f"={BENEISH_FLAG}", None, color=C_TEXT)
Dh["C12"].number_format = "General"
lbl(Dh, "B13", "Piotroski")
val(Dh, "C13", f'={PIOTROSKI}&" / 9"', None, color=C_TEXT)
Dh["C13"].number_format = "General"

# ---- VERDICT CHIP (big) -- F6, driven by MoS vs quality-tiered required MoS (C24) + risk gates
Dh.merge_cells("F6:H8")
verdict_formula = (
    f'=IF(OR({ALTMAN}<1.81,{BENEISH}>-1.78),"HOLD (RISK FLAG)",'
    f'IF(C10>=$C$24+0.15,"STRONG BUY",'
    f'IF(C10>=$C$24,"BUY - MEETS BAR",'
    f'IF(C10>=$C$24-0.15,"HOLD / WATCH","OVERVALUED / AVOID"))))'
)
vc = Dh["F6"]
vc.value = verdict_formula
vc.font = Font(name=F_UI, size=20, bold=True, color=C_BG)
vc.alignment = Alignment(horizontal="center", vertical="center")
vc.fill = fill_amber
# CF by MoS (C10) vs required bar (C24) + risk; local cells only. Risk rule added last = top priority.
Dh.conditional_formatting.add("F6",
    FormulaRule(formula=['AND($C$10>=$C$24+0.15,$C$18>=1.81,$C$19<=-1.78)'],
                fill=fill_green, font=Font(name=F_UI, size=20, bold=True, color=C_BG)))
Dh.conditional_formatting.add("F6",
    FormulaRule(formula=['AND($C$10>=$C$24,$C$10<$C$24+0.15)'],
                fill=fill_teal, font=Font(name=F_UI, size=20, bold=True, color=C_BG)))
Dh.conditional_formatting.add("F6",
    FormulaRule(formula=['AND($C$10>=$C$24-0.15,$C$10<$C$24)'],
                fill=fill_amber, font=Font(name=F_UI, size=20, bold=True, color=C_BG)))
Dh.conditional_formatting.add("F6",
    FormulaRule(formula=['$C$10<$C$24-0.15'],
                fill=fill_red, font=Font(name=F_UI, size=20, bold=True, color=C_TEXT)))
Dh.conditional_formatting.add("F6",
    FormulaRule(formula=['OR($C$18<1.81,$C$19>-1.78)'],
                fill=fill_red, font=Font(name=F_UI, size=20, bold=True, color=C_TEXT)))

# Entry zone block
section_label(Dh, "F10", "ENTRY ZONE", fill=fill_raised)
for col in "GH":
    Dh[f"{col}10"].fill = fill_raised
lbl(Dh, "F11", "Buy below (FV*(1-req MoS))", size=10, font_color=C_TEXT)
val(Dh, "G11", "=C8*(1-C24)", NF_USD, color=C_GOLD_BR, bold=True, size=12)
lbl(Dh, "F12", "Buy band low (FV*(1-req-10%))")
val(Dh, "G12", "=C8*(1-C24-0.10)", NF_USD, color=C_GREEN)
lbl(Dh, "F13", "Buy band high (= buy below)")
val(Dh, "G13", "=C8*(1-C24)", NF_USD, color=C_GREEN)
lbl(Dh, "F14", "Nearest Fib support")
val(Dh, "G14", f"={NEAR_FIB}", NF_USD, color=C_BLUE, bold=True)

# Score chips block
section_label(Dh, "B16", "SCORE CHIPS", fill=fill_raised)
for col in "CD":
    Dh[f"{col}16"].fill = fill_raised
lbl(Dh, "B17", "Piotroski F-Score")
val(Dh, "C17", f"={PIOTROSKI}", "0", color=C_TEXT, bold=True)
Dh.conditional_formatting.add("C17",
    CellIsRule(operator="greaterThanOrEqual", formula=["7"], fill=fill_green, font=Font(name=F_MONO,bold=True,color=C_BG)))
Dh.conditional_formatting.add("C17",
    CellIsRule(operator="between", formula=["4","6"], fill=fill_amber, font=Font(name=F_MONO,bold=True,color=C_BG)))
Dh.conditional_formatting.add("C17",
    CellIsRule(operator="lessThan", formula=["4"], fill=fill_red, font=Font(name=F_MONO,bold=True,color=C_TEXT)))
lbl(Dh, "B18", "Altman Z")
val(Dh, "C18", f"={ALTMAN}", "0.00", color=C_TEXT, bold=True)
Dh.conditional_formatting.add("C18",
    CellIsRule(operator="greaterThan", formula=["2.99"], fill=fill_green, font=Font(name=F_MONO,bold=True,color=C_BG)))
Dh.conditional_formatting.add("C18",
    CellIsRule(operator="between", formula=["1.81","2.99"], fill=fill_amber, font=Font(name=F_MONO,bold=True,color=C_BG)))
Dh.conditional_formatting.add("C18",
    CellIsRule(operator="lessThan", formula=["1.81"], fill=fill_red, font=Font(name=F_MONO,bold=True,color=C_TEXT)))
lbl(Dh, "B19", "Beneish M")
val(Dh, "C19", f"={BENEISH}", "0.000", color=C_TEXT, bold=True)
Dh.conditional_formatting.add("C19",
    CellIsRule(operator="greaterThan", formula=["-1.78"], fill=fill_red, font=Font(name=F_MONO,bold=True,color=C_TEXT)))
Dh.conditional_formatting.add("C19",
    CellIsRule(operator="lessThanOrEqual", formula=["-1.78"], fill=fill_green, font=Font(name=F_MONO,bold=True,color=C_BG)))
lbl(Dh, "B20", "ROIC vs WACC spread")
val(Dh, "C20", f"={ROIC}-{WACC}", NF_PCT2, color=C_GREEN, bold=True)
Dh.conditional_formatting.add("C20",
    CellIsRule(operator="greaterThan", formula=["0"], fill=fill_green, font=Font(name=F_MONO,bold=True,color=C_BG)))
Dh.conditional_formatting.add("C20",
    CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=fill_red, font=Font(name=F_MONO,bold=True,color=C_TEXT)))

# ---- CONTEXT & REGIME chips (C24 = effective required MoS, read by the verdict) ----
section_label(Dh, "B22", "CONTEXT & REGIME", fill=fill_raised)
for col in "CD":
    Dh[f"{col}22"].fill = fill_raised
lbl(Dh, "B23", "Quality tier")
val(Dh, "C23", f"={QUAL_TIER}", None, color=C_TEXT, bold=True); Dh["C23"].number_format = "General"
lbl(Dh, "B24", "Required MoS (effective)")
val(Dh, "C24", f"={REQ_MOS_EFF}", NF_PCT, color=C_GOLD_BR, bold=True)
lbl(Dh, "B25", "Valuation vs own 10y")
val(Dh, "C25", f"={VALHIST_LABEL}", None, color=C_TEXT, bold=True); Dh["C25"].number_format = "General"
lbl(Dh, "B26", "Market regime")
val(Dh, "C26", f"={REGIME_LABEL}", None, color=C_TEXT, bold=True); Dh["C26"].number_format = "General"
lbl(Dh, "B27", "Meets bar to buy?")
val(Dh, "C27", '=IF(C10>=C24,"YES","NO")', None, color=C_GREEN, bold=True); Dh["C27"].number_format = "General"

# ---- VALUATION SUMMARY TABLE ----
section_label(Dh, "F16", "VALUATION SUMMARY", fill=fill_raised)
for col in "GH":
    Dh[f"{col}16"].fill = fill_raised
style_cell(Dh, "F17", "Method", font_color=C_MUTED, size=8, bold=True, fill=fill_raised)
style_cell(Dh, "G17", "$/sh", font_color=C_MUTED, size=8, bold=True, fill=fill_raised, align="right")
style_cell(Dh, "H17", "Upside", font_color=C_MUTED, size=8, bold=True, fill=fill_raised, align="right")
summary_methods = [
    ("DCF (FCFF)", DCF_FCFF_PS),
    ("DCF (FCFE)", DCF_FCFE_PS),
    ("Relative (P/E & EV/EBITDA)", REL_BLEND),
    ("EPV", EPV_PS),
    ("Graham Number", GRAHAM_NUM),
    ("Graham Formula", GRAHAM_FORM),
    ("Prob-weighted scenario", SC_PW),
    ("Monte Carlo median", MC_P50),
]
sm0 = 18
for i,(name,ref) in enumerate(summary_methods):
    r = sm0 + i
    lbl(Dh, f"F{r}", name, size=9, font_color=C_TEXT)
    val(Dh, f"G{r}", f"=IFERROR({ref},NA())", NF_USD, color=C_TEXT)
    val(Dh, f"H{r}", f"=IFERROR({ref}/C6-1,NA())", NF_PCT, color=C_MUTED)
sm_last = sm0 + len(summary_methods) - 1
# blended row
br = sm_last + 1
lbl(Dh, f"F{br}", "BLENDED FAIR VALUE", size=10, font_color=C_GOLD_BR)
val(Dh, f"G{br}", "=C8", NF_USD, color=C_GOLD_BR, bold=True)
val(Dh, f"H{br}", "=C8/C6-1", NF_PCT, color=C_GREEN, bold=True)

# ---- one-line thesis ----
thesis_row = br + 2
section_label(Dh, f"B{thesis_row}", "THESIS", fill=fill_raised)
for col in "CDEFGH":
    Dh[f"{col}{thesis_row}"].fill = fill_raised
Dh.merge_cells(f"B{thesis_row+1}:H{thesis_row+2}")
thesis = (
    f'="At $"&TEXT(C6,"0.00")&", APEX trades at "&TEXT(C9,"0.0%")&'
    f'" upside to blended FV of $"&TEXT(C8,"0.00")&" (MoS "&TEXT(C10,"0.0%")&"). "&'
    f'"ROIC "&IF({ROIC}>{WACC},"exceeds","trails")&" WACC; Altman "&{ALTMAN_ZONE}&'
    f'", Beneish "&{BENEISH_FLAG}&", Piotroski "&{PIOTROSKI}&"/9. Buy below $"&TEXT(G11,"0.00")&"."'
)
tc = Dh[f"B{thesis_row+1}"]
tc.value = thesis
tc.font = Font(name=F_UI, size=10, italic=True, color=C_TEXT)
tc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
tc.fill = fill_panel

# Caution notes (Piotroski<4) under thesis
note_row = thesis_row + 3
nc = Dh[f"B{note_row}"]
Dh.merge_cells(f"B{note_row}:H{note_row}")
nc.value = (f'=IF({PIOTROSKI}<4,"⚠ Caution: Piotroski below 4 — weak fundamental momentum.",'
            f'IF(OR({ALTMAN}<1.81,{BENEISH}>-1.78),"⚠ Risk flag active — verdict capped at HOLD.",""))')
nc.font = Font(name=F_UI, size=9, italic=True, color=C_RED)
nc.alignment = Alignment(horizontal="left", vertical="center")
nc.fill = fill_panel

# ============================================================================
# WORKBOOK CALCULATION SETTINGS
# ============================================================================
wb.calculation.calcMode = "auto"
wb.calculation.fullCalcOnLoad = True

# Save
OUT = "Equity_Valuation_Cockpit_AAPL.xlsx"
wb.save(OUT)
print("SAVED:", OUT)
